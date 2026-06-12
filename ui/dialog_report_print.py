"""
报告打印对话框模块 - 报告打印前的编辑确认对话框和 XLSX 生成

本模块提供以下主要功能：
  1. show_report_dialog 函数 - 报告打印前的编辑确认对话框（14 个可编辑字段）
  2. _create_report_xlsx_data 函数 - 根据编辑框提交的数据创建测评报告打印信息 XLSX
  3. _create_report_xlsx 函数 - 创建 XLSX 基础结构（表头、固定数据行、附件超链接）

特色功能：
  - 支持设置默认值（保存到 data/report_defaults.json）
  - 空值自动回退到项目已有数据兜底
  - 可滚动内容区域（Canvas + Scrollbar）
  - 模态窗口（grab_set() 禁止操作父窗口）
  - XLSX 生成包含 21 列表头和附件超链接

依赖关系：
  - models.Project：项目数据实体类
  - utils.Config：全局 UI 配置（字体、颜色、尺寸等）
  - utils.helpers.bordered_entry：边框输入框工厂函数（用于报告对话框）
  - openpyxl：Excel 文件读写（用于生成报告打印信息表）
"""

import tkinter as tk  # Python 标准 GUI 库，提供 Toplevel、Label、Button、Frame、Canvas 等组件
import json, os  # JSON 解析库和文件系统操作库
from datetime import date  # 日期类，用于截止日期的计算和比较

from utils.config import Config  # 全局配置类：提供字体、颜色、尺寸预警天数等 UI 常量
from utils.helpers import bordered_entry  # 带边框的输入框工厂函数


# =============================================================================
# show_report_dialog 函数 - 报告打印编辑确认对话框（公开接口，原 _show_report_dialog）
# =============================================================================

def show_report_dialog(parent, cname="", sname="", location="", deadline=""):
    """显示报告打印信息的编辑确认对话框

    此对话框用于在打印测评报告前，让用户确认和编辑以下 14 个字段：
      1. 客户公司全称
      2. 合同编号或项目名称
      3. 所属地
      4. 系统名称
      5. 是否录入 CRM
      6. 编制/审核/批准日期
      7. 编制人
      8. 审核人
      9. 渗透人员
      10. 测评结论及重大风险隐患数量
      11. 盖章
      12. 打印要求
      13. 项目组长联系人
      14. 实际报告编制人

    特色功能：
      - 支持设置默认值（保存到 data/report_defaults.json）
      - 空值自动回退到项目已有数据兜底
      - 可滚动内容区域（Canvas + Scrollbar）
      - 模态窗口（grab_set() 禁止操作父窗口）

    Args:
        parent: 父级窗口（用于定位和模态绑定）
        cname: 预填的客户公司名称（来自 project.company_name）
        sname: 预填的系统名称（来自 project.system_name）
        location: 预填的所属地（来自 project.location，取省/市前缀）
        deadline: 预填的日期（来自 project.deadline，或当天日期）

    Returns:
        dict | None: 用户确认后返回 14 个字段的字典，取消返回 None
    """
    from tkinter import messagebox  # 消息弹窗，用于保存默认值后的提示

    # ---- 创建对话框主窗口 ----
    dlg = tk.Toplevel(parent)  # 创建子顶级窗口
    dlg.title("报告打印信息确认")  # 设置窗口标题
    dlg.geometry("520x680")  # 设置初始窗口尺寸
    dlg.minsize(440, 500)  # 设置最小尺寸限制
    dlg.configure(bg="#ffffff")  # 白色背景
    dlg.grab_set()  # 设置为模态窗口（禁止操作父窗口）
    dlg.resizable(True, True)  # 允许用户自由调整窗口大小

    # 用于存储用户确认结果的闭包变量
    result = {"confirmed": False}  # 初始为未确认状态

    # ---- 加载保存的默认值 ----
    defaults = {}  # 存储从文件加载的默认值
    def_path = os.path.join(Config.get_data_dir(), "data", "report_defaults.json")  # 默认值文件路径
    if os.path.exists(def_path):  # 如果文件存在
        try:
            with open(def_path, "r", encoding="utf-8") as f:  # 以 UTF-8 编码打开
                defaults = json.load(f)  # 加载 JSON 内容到字典
        except Exception:  # 任何解析异常（JSON 格式错误、权限不足等）
            pass  # 静默处理：保持空字典，不影响正常使用

    # ========== 底部按钮区域（先 pack 以确保障空间） ==========
    btn_frame = tk.Frame(dlg, bg="#f0f2f5")  # 按钮栏容器 Frame，浅灰背景
    btn_frame.pack(fill=tk.X, side=tk.BOTTOM)  # 固定在对话框底部，水平填充
    tk.Frame(btn_frame, bg="#d0d5dd", height=1).pack(fill=tk.X)  # 1px 灰色分割线
    btn_inner = tk.Frame(btn_frame, bg="#f0f2f5")  # 按钮内容区域（提供内边距）
    btn_inner.pack(fill=tk.X, padx=16, pady=8)  # 水平填充，16px 水平 + 8px 垂直边距

    # ========== 可滚动内容区域 ==========
    canvas = tk.Canvas(dlg, bg="#ffffff", highlightthickness=0)  # Canvas 滚动容器，白色背景，无高亮
    scrollbar = tk.Scrollbar(dlg, orient=tk.VERTICAL, command=canvas.yview)  # 垂直滚动条
    canvas.configure(yscrollcommand=scrollbar.set)  # 绑定 Canvas 与滚动条双向联动
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)  # 滚动条固定在右侧，垂直填充
    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)  # Canvas 填充左侧剩余空间

    main = tk.Frame(canvas, bg="#ffffff", padx=20, pady=15)  # 内容主容器，白色背景，20px 水平 + 15px 垂直内边距
    cw = canvas.create_window((0, 0), window=main, anchor="nw")  # 在 Canvas 中创建窗口（嵌入 main Frame）
    main.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))  # 内容变化时更新滚动区域
    canvas.bind("<Configure>", lambda e: canvas.itemconfig(cw, width=e.width))  # Canvas 宽度变化时同步窗口宽度
    # 鼠标滚轮绑定（Canvas 和对话框都支持滚轮滚动）
    canvas.bind("<MouseWheel>", lambda e: canvas.yview_scroll(int(-e.delta/120), "units"))
    dlg.bind("<MouseWheel>", lambda e: canvas.yview_scroll(int(-e.delta/120), "units"))

    # ========== 创建表单字段的工具函数 ==========
    def _make_row(label, default=""):
        tk.Label(main, text=label, bg="#ffffff",
                 font=(Config.FONT_FAMILY, Config.FONT_SIZE_NORMAL)).pack(anchor="w")
        var = tk.StringVar(value=default)
        _, outer = bordered_entry(main, textvariable=var)
        outer.pack(fill=tk.X, pady=(2, 6))
        return var

    def _make_text_row(label, default="", height=3):
        """创建多行文本框：标签 + 带边框的 Text 组件，用于多系统字段。"""
        tk.Label(main, text=label, bg="#ffffff",
                 font=(Config.FONT_FAMILY, Config.FONT_SIZE_NORMAL)).pack(anchor="w")
        outer = tk.Frame(main, bg="#d0d5dd")
        inner = tk.Frame(outer, bg="#ffffff")
        inner.pack(fill=tk.BOTH, expand=True, padx=1, pady=1)
        text = tk.Text(inner, font=("Microsoft YaHei", 10), relief="flat",
                       borderwidth=0, height=height, wrap="word")
        text.pack(fill=tk.BOTH, expand=True)
        text.insert("1.0", default)
        outer.pack(fill=tk.X, pady=(2, 6))
        return text

    # ========== 对话框标题 ==========
    tk.Label(main, text="报告打印信息确认", bg="#ffffff", fg="#2c3e50",  # 深色标题文字
             font=(Config.FONT_FAMILY, Config.FONT_SIZE_HEADER, "bold"),  # 大号粗体
             ).pack(anchor="w", pady=(0, 10))  # 左对齐，下方 10px 间距

    # ========== 创建 14 个字段输入行 ==========
    # 优先使用已保存的默认值，否则使用项目数据；两者都无则留空
    v_cname = _make_row("客户公司全称", defaults.get("cname") or cname)
    v_contract = _make_row("合同编号或项目名称", defaults.get("contract") or f"{cname}网络安全等级保护测评服务项目")
    v_location = _make_row("所属地", defaults.get("location") or location)
    t_sname = _make_text_row("系统名称", defaults.get("sname") or sname, height=3)
    v_crm = _make_row("是否录入CRM", defaults.get("crm") or "是")  # 默认"是"
    v_deadline = _make_row("编制/审核/批准日期", defaults.get("deadline") or deadline)
    v_author = _make_row("编制人", defaults.get("author") or "")
    v_reviewer = _make_row("审核人", defaults.get("reviewer") or "")
    v_pentester = _make_row("渗透人员", defaults.get("pentester") or "")
    v_conclusion = _make_row("测评结论及重大风险隐患数量", defaults.get("conclusion") or "")
    v_seal = _make_row("盖章", defaults.get("seal") or "")
    v_print_req = _make_row("打印要求", defaults.get("print_req") or "")
    v_leader = _make_row("项目组长联系人", defaults.get("leader") or "")
    v_actual = _make_row("实际报告编制人", defaults.get("actual_author") or "")

    # ========== 辅助函数：收集所有字段的值 ==========
    def _collect():
        """收集 14 个字段的当前值并返回字典

        Returns:
            dict: 包含所有字段名称-值映射的字典
        """
        return {
            "cname": v_cname.get().strip(),
            "contract": v_contract.get().strip(),
            "location": v_location.get().strip(),
            "sname": t_sname.get("1.0", "end-1c").strip(),
            "crm": v_crm.get().strip(),
            "deadline": v_deadline.get().strip(),
            "author": v_author.get().strip(),
            "reviewer": v_reviewer.get().strip(),
            "pentester": v_pentester.get().strip(),
            "conclusion": v_conclusion.get().strip(),
            "seal": v_seal.get().strip(),
            "print_req": v_print_req.get().strip(),
            "leader": v_leader.get().strip(),
            "actual_author": v_actual.get().strip(),
        }

    # ========== 确认按钮的回调函数 ==========
    def _on_confirm():
        """用户点击"确认"按钮的处理

        收集所有字段值，对空字段进行兜底填充（使用项目数据），
        设置 confirmed = True 并关闭对话框。
        """
        data = _collect()  # 收集所有字段的当前值

        # --- 空值回退机制：优先用户输入，否则用项目数据兜底 ---
        if not data["cname"]: data["cname"] = cname  # 公司名称回退
        if not data["location"]: data["location"] = location  # 所属地回退
        if not data["sname"]: data["sname"] = sname  # 系统名称回退
        if not data["deadline"]: data["deadline"] = deadline  # 日期回退
        if not data["contract"]: data["contract"] = f"{cname}网络安全等级保护测评服务项目"  # 合同回退

        result["confirmed"] = True  # 标记为已确认
        result["data"] = data  # 存储收集到的数据
        dlg.destroy()  # 关闭对话框

    # ========== 设置默认值按钮的回调函数 ==========
    def _open_defaults_editor():
        """打开设置默认值的独立编辑窗口

        功能：
          1. 读取已保存的默认值文件（data/report_defaults.json）
          2. 如果无保存值，使用当前对话框中的数据
          3. 允许用户编辑所有 14 个字段的默认值
          4. 保存到 JSON 文件
          5. 回填到主对话框的对应字段
        """
        # 加载已有默认值
        path = os.path.join(Config.get_data_dir(), "data", "report_defaults.json")
        saved = {}  # 存储已保存的默认值
        if os.path.exists(path):  # 文件存在
            try:
                with open(path, "r", encoding="utf-8") as f:
                    saved = json.load(f)  # 加载 JSON
            except Exception:
                pass  # 解析失败则保持空字典

        # 如果没有任何保存值，用当前对话框的数据预填
        if not any(saved.values()):
            saved = _collect()

        # ---- 创建设置默认值的编辑窗口 ----
        dedit = tk.Toplevel(dlg)  # 子窗口（父窗口为 dlg）
        dedit.title("设置默认值")
        dedit.geometry("500x600")  # 初始尺寸
        dedit.minsize(420, 400)  # 最小尺寸限制
        dedit.configure(bg="#ffffff")  # 白色背景
        dedit.grab_set()  # 模态窗口

        # --- 底部按钮区域（先 pack 以确保障空间） ---
        dbtn_frame = tk.Frame(dedit, bg="#f0f2f5")  # 按钮容器
        dbtn_frame.pack(fill=tk.X, side=tk.BOTTOM)  # 固定在底部
        tk.Frame(dbtn_frame, bg="#d0d5dd", height=1).pack(fill=tk.X)  # 1px 分割线
        dbtn_inner = tk.Frame(dbtn_frame, bg="#f0f2f5")  # 按钮内容区域
        dbtn_inner.pack(fill=tk.X, padx=16, pady=8)

        # --- 可滚动内容区域 ---
        dcanvas = tk.Canvas(dedit, bg="#ffffff", highlightthickness=0)  # Canvas 容器
        dscroll = tk.Scrollbar(dedit, orient=tk.VERTICAL, command=dcanvas.yview)  # 滚动条
        dcanvas.configure(yscrollcommand=dscroll.set)  # 双向联动
        dscroll.pack(side=tk.RIGHT, fill=tk.Y)
        dcanvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        dmain = tk.Frame(dcanvas, bg="#ffffff", padx=20, pady=15)  # 内容主容器
        dw = dcanvas.create_window((0, 0), window=dmain, anchor="nw")  # 嵌入 Canvas
        dmain.bind("<Configure>", lambda e: dcanvas.configure(scrollregion=dcanvas.bbox("all")))
        dcanvas.bind("<Configure>", lambda e: dcanvas.itemconfig(dw, width=e.width))
        # 滚轮支持
        dcanvas.bind("<MouseWheel>", lambda e: dcanvas.yview_scroll(int(-e.delta/120), "units"))
        dedit.bind("<MouseWheel>", lambda e: dcanvas.yview_scroll(int(-e.delta/120), "units"))

        # --- 创建 14 个字段的输入行 ---
        text_keys = {"sname"}  # 多行文本框字段
        dvars = {}
        for label, key in [
            ("客户公司全称", "cname"), ("合同编号或项目名称", "contract"),
            ("所属地", "location"), ("系统名称", "sname"), ("是否录入CRM", "crm"),
            ("编制/审核/批准日期", "deadline"), ("编制人", "author"),
            ("审核人", "reviewer"), ("渗透人员", "pentester"),
            ("测评结论及重大风险隐患数量", "conclusion"), ("盖章", "seal"),
            ("打印要求", "print_req"), ("项目组长联系人", "leader"),
            ("实际报告编制人", "actual_author"),
        ]:
            tk.Label(dmain, text=label, bg="#ffffff",
                     font=(Config.FONT_FAMILY, Config.FONT_SIZE_NORMAL)).pack(anchor="w")
            if key in text_keys:
                outer = tk.Frame(dmain, bg="#d0d5dd")
                inner = tk.Frame(outer, bg="#ffffff")
                inner.pack(fill=tk.BOTH, expand=True, padx=1, pady=1)
                w = tk.Text(inner, font=("Microsoft YaHei", 10), relief="flat",
                            borderwidth=0, height=3, wrap="word", width=40)
                w.pack(fill=tk.BOTH, expand=True)
                w.insert("1.0", saved.get(key) or "")
                outer.pack(fill=tk.X, pady=(2, 5))
                dvars[key] = w
            else:
                var = tk.StringVar(value=saved.get(key) or "")
                _, outer = bordered_entry(dmain, textvariable=var, width=40)
                outer.pack(fill=tk.X, pady=(2, 5))
                dvars[key] = var

        def _save_and_close():
            data = {}
            for k, v in dvars.items():
                if isinstance(v, tk.Text):
                    data[k] = v.get("1.0", "end-1c").strip()
                else:
                    data[k] = v.get().strip()
            os.makedirs(os.path.dirname(path), exist_ok=True)
            with open(path, "w", encoding="utf-8") as f:
                json.dump(data, f, ensure_ascii=False, indent=2)

            # 回填到主对话框
            for key, val in dvars.items():
                target_map = {
                    "cname": v_cname, "contract": v_contract, "location": v_location,
                    "sname": t_sname, "crm": v_crm, "deadline": v_deadline,
                    "author": v_author, "reviewer": v_reviewer, "pentester": v_pentester,
                    "conclusion": v_conclusion, "seal": v_seal,
                    "print_req": v_print_req, "leader": v_leader,
                    "actual_author": v_actual,
                }
                target = target_map.get(key)
                if target is None:
                    continue
                if isinstance(target, tk.Text):
                    target.delete("1.0", "end")
                    target.insert("1.0", data[key])
                else:
                    target.set(data[key])

            messagebox.showinfo("提示", "默认值已保存", parent=dedit)
            dedit.destroy()

        # --- 编辑窗口的按钮 ---
        tk.Button(dbtn_inner, text="取消", command=dedit.destroy,
                  bg="#ffffff", fg="#2c3e50", cursor="hand2",
                  font=(Config.FONT_FAMILY, Config.FONT_SIZE_SMALL),
                  relief="flat", padx=20, pady=5,
                  highlightbackground="#d0d5dd", highlightthickness=1,
                  ).pack(side=tk.RIGHT, padx=(10, 0))  # "取消"按钮，右侧
        tk.Button(dbtn_inner, text="保存", command=_save_and_close,
                  bg="#3498db", fg="white", cursor="hand2",
                  font=(Config.FONT_FAMILY, Config.FONT_SIZE_NORMAL, "bold"),
                  relief="flat", padx=20, pady=5,
                  ).pack(side=tk.RIGHT)  # "保存"按钮，右侧
        # 快捷键：Enter 保存，Escape 取消
        dedit.bind("<Return>", lambda e: _save_and_close())
        dedit.bind("<Escape>", lambda e: dedit.destroy())

    # ========== 主对话框的按钮 ==========
    # "设置默认值"按钮（左侧）
    tk.Button(btn_inner, text="设置默认值", command=_open_defaults_editor,
              bg="#f0f2f5", fg="#2c3e50", cursor="hand2",
              font=(Config.FONT_FAMILY, Config.FONT_SIZE_SMALL),
              relief="flat", padx=12, pady=5,
              activebackground="#d5dbdb",  # 按下时的背景色
              ).pack(side=tk.LEFT)

    # "取消"按钮（右侧）
    tk.Button(btn_inner, text="取消", command=dlg.destroy,  # 直接关闭对话框
              bg="#ffffff", fg="#2c3e50", cursor="hand2",
              font=(Config.FONT_FAMILY, Config.FONT_SIZE_SMALL),
              relief="flat", padx=20, pady=5,
              highlightbackground="#d0d5dd", highlightthickness=1,  # 灰色边框
              activebackground="#f0f2f5",
              ).pack(side=tk.RIGHT, padx=(10, 0))

    # "确认"按钮（右侧）
    tk.Button(btn_inner, text="确认", command=_on_confirm,
              bg="#3498db", fg="white", cursor="hand2",  # 蓝色主题
              font=(Config.FONT_FAMILY, Config.FONT_SIZE_NORMAL, "bold"),  # 粗体
              relief="flat", padx=20, pady=5,
              activebackground="#2980b9",  # 按下时的背景色（深蓝）
              ).pack(side=tk.RIGHT)

    # ========== 快捷键绑定 ==========
    dlg.bind("<Return>", lambda e: _on_confirm())  # Enter 键 -> 确认
    dlg.bind("<Escape>", lambda e: dlg.destroy())  # Escape 键 -> 取消（关闭）

    # ========== 等待对话框关闭 ==========
    parent.wait_window(dlg)  # 阻塞等待对话框关闭（模态行为）

    # 返回结果
    if result["confirmed"]:  # 用户点击了确认
        return result["data"]  # 返回 14 个字段的字典
    return None  # 用户取消，返回 None


# =============================================================================
# _create_report_xlsx_data / _create_report_xlsx - XLSX 生成
# =============================================================================

def _create_report_xlsx_data(project, path, data, report_dir, root,
                             all_projects: list = None):
    """根据编辑框提交的数据创建测评报告打印信息 XLSX 文件

    这是两步 XLSX 生成流程的入口：
      1. 先调用 _create_report_xlsx 创建基础 XLSX（含表头、固定数据行）
      2. 再用编辑框的额外数据覆盖 B3、D3、J3~U3 等单元格

    Args:
        project: 项目实体对象（用于获取 deadline、location 等字段）
        path: XLSX 文件保存路径
        data: 用户确认的 14 个字段字典（来自 show_report_dialog 返回值）
        report_dir: 报告打印目录路径（用于附件超链接）
        root: 项目根目录路径（用于附件超链接查找）
        all_projects: 合并卡片的所有项目（用于判断多系统）
    """
    wb, ws, ole_entries = _create_report_xlsx(project, data["cname"], data["sname"],
                                                report_dir, root, all_projects=all_projects)

    from openpyxl.styles import Alignment, Border, Side

    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin'))

    is_multi = all_projects and len(all_projects) > 1
    num_rows = len(all_projects) if is_multi else 1

    shared_cols = {
        'B': data.get("contract", ""),
        'D': data.get("crm", "是"),
        'J': data.get("author", ""),
        'K': data.get("reviewer", ""),
        'L': data.get("pentester", ""),
        'M': data.get("conclusion", ""),
        'N': data.get("seal", ""),
        'S': data.get("print_req", ""),
        'T': data.get("leader", ""),
        'U': data.get("actual_author", ""),
    }
    for row_idx in range(num_rows):
        row_num = 3 + row_idx
        for col, val in shared_cols.items():
            cell = ws[f'{col}{row_num}']
            cell.value = val
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    wb.save(path)
    wb.close()

    if ole_entries:
        _embed_oles_in_xlsx(path, ole_entries)


def _embed_oles_in_xlsx(xlsx_path, ole_entries):
    """通过 OLE 嵌入服务将附件文件作为 OLE 对象嵌入 XLSX 文件。

    使用 Win32 COM 接口（Win32ComOleEmbedService）将指定的文件列表
    作为嵌入式 OLE 对象链接到 XLSX 的指定单元格位置。

    该功能属于技术隔离层（原则 #5），OLE 服务不可用时静默跳过，
    不影响 XLSX 文件本身的生成——此时单元格仍显示附件文件名文本。

    Args:
        xlsx_path: 已生成并保存的 XLSX 文件完整路径
        ole_entries: OLE 嵌入条目列表，每项为 (列字母, 行号, 文件路径) 元组，
            例如 [("O", 3, "C:\\reports\\基本情况表.docx"), ...]

    Returns:
        None: OLE 嵌入失败不阻塞报告打印主流程，静默返回
    """
    from services.ole_service import Win32ComOleEmbedService
    from services.interfaces import OleEmbedError

    svc = Win32ComOleEmbedService()
    if not svc.is_available():
        return  # OLE 服务不可用时静默跳过，表格仍有文件名

    try:
        svc.embed_files(xlsx_path, ole_entries)
    except OleEmbedError:
        pass  # 嵌入失败不阻塞报告打印主流程


def _create_report_xlsx(project, cname, sname, report_dir, root,
                        all_projects: list = None):
    """创建测评报告打印信息 XLSX 的基础结构

    多系统: 附件搜索覆盖根目录 + 各系统子目录。

    表头列（A~U 共 21 列）：
      A: 序号 | B: 合同编号 | C: 客户公司 | D: 是否录入CRM | E: 所属地
      F: 系统名称 | G: 编制日期 | H: 审核日期 | I: 批准日期
      J: 编制人 | K: 审核人 | L: 渗透人员 | M: 测评结论数量 | N: 盖章
      O: 基本情况表附件 | P: 授权书附件 | Q: 报告附件 | R: 归档附件
      S: 打印要求 | T: 项目组长 | U: 实际编制人

    Args:
        project: 项目实体对象
        path: XLSX 文件保存路径
        cname: 客户公司全称
        sname: 系统名称
        report_dir: 报告打印目录（优先搜索附件）
        root: 项目根目录（兜底搜索附件）
        all_projects: 合并卡片的所有项目（用于判断多系统）
    """
    import os as _os  # 操作系统模块（使用别名避免与外部 os 冲突）
    import openpyxl  # Excel 文件操作库
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill  # Excel 样式组件
    from datetime import date  # 当前日期

    # ---- 创建 Workbook 和默认工作表 ----
    wb = openpyxl.Workbook()  # 创建新的 Excel 工作簿
    ws = wb.active  # 获取默认的活动工作表
    ws.title = "Sheet1"  # 设置工作表名称为 "Sheet1"

    # ---- 设置列宽（A~U 共 21 列） ----
    widths = {'A': 6, 'B': 30, 'C': 22, 'D': 10, 'E': 10, 'F': 18,
              'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 10, 'L': 10,
              'M': 16, 'N': 10, 'O': 16, 'P': 16, 'Q': 14, 'R': 14,
              'S': 16, 'T': 10, 'U': 14}
    for col, w in widths.items():  # 遍历列宽设置
        ws.column_dimensions[col].width = w  # 设置每列宽度

    # ---- 提示行 (第 1 行) ----
    ws['G1'] = '（需要签入报告，请确认）'  # 红色提醒文字
    ws['G1'].font = Font(color='FF0000', bold=True)  # 红色粗体

    # ---- 表头 (第 2 行) ----
    headers = ['序号', '合同编号或项目名称', '客户公司全称', '是否录入CRM', '所属地',
               '系统名称', '编制日期', '审核日期', '批准日期',
               '编制人（与联盟系统对应）', '审核人', '渗透人员',
               '测评结论及重大风险隐患数量', '盖章',
               '等级测评项目基本情况表附件', '授权书及风险告知书附件',
               '测评报告附件', '测评归档附件', '打印要求',
               '项目组长联系人', '实际报告编制人']  # 21 个表头

    # 表头样式定义
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')  # 浅蓝填充
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),  # 左右细线
        top=Side(style='thin'), bottom=Side(style='thin'))  # 上下细线

    for i, h in enumerate(headers, 1):  # 遍历表头，i 从 1（Excel 列号从 1 开始）
        cell = ws.cell(row=2, column=i, value=h)  # 在第 2 行写入表头
        cell.font = Font(bold=True, size=10)  # 粗体 10 号
        cell.fill = header_fill  # 浅蓝背景填充
        cell.border = thin_border  # 细线边框
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # 居中自动换行
    ws.row_dimensions[2].height = 35  # 表头行高 35px

    # ---- 数据行 ----
    is_multi = all_projects and len(all_projects) > 1
    date_val = project.deadline or date.today().strftime('%Y-%m-%d')
    num_rows = len(all_projects) if is_multi else 1

    for row_idx in range(num_rows):
        row_num = 3 + row_idx
        cur = all_projects[row_idx] if is_multi else project

        sys_loc = (cur.location or "").split("-")[0] if cur.location else ""
        sys_name = cur.system_name or ""

        row_data = [
            (row_idx + 1, 'A'),
            (f'{cname}网络安全等级保护测评服务项目', 'B'),
            (cname, 'C'),
            ('是', 'D'),
            (sys_loc, 'E'),
            (sys_name, 'F'),
            (date_val, 'G'), (date_val, 'H'), (date_val, 'I'),
            ('双击打开附件', 'O'), ('双击打开附件', 'P'),
            ('双击打开附件', 'Q'), ('双击打开附件', 'R'),
        ]
        for val, col in row_data:
            cell = ws[f'{col}{row_num}']
            cell.value = val
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        for col_idx in range(1, 22):
            cell = ws.cell(row=row_num, column=col_idx)
            cell.border = thin_border
            if not cell.value:
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.row_dimensions[row_num].height = 25

    # ---- 合并单元格 G1-U1（提示行合并为一整行） ----
    ws.merge_cells('G1:U1')  # 将 G1 到 U1 合并为一个单元格

    # ---- 附件搜索与 OLE 嵌入 ----
    link_map = {
        'O': ['基本情况表'],
        'P': ['测评授权书', '风险告知书', '放弃工具测试声明'],
        'Q': ['测评报告-终稿'],
        'R': ['过程文档.zip'],
    }
    # Q 列限定 PDF / R 列限定 ZIP
    ext_filter = {
        'Q': lambda fn: fn.lower().endswith('.pdf'),
        'R': lambda fn: fn.lower().endswith('.zip'),
    }

    # 收集系统子目录（多系统时用于按系统优先匹配附件）
    sys_subdirs = []
    if is_multi:
        for dname in _os.listdir(root):
            dpath = _os.path.join(root, dname)
            if _os.path.isdir(dpath) and "报告打印" not in dname \
                    and dname != "01-其他归档文件":
                sys_subdirs.append(dpath)

    ole_entries = []

    for row_idx in range(num_rows):
        row_num = 3 + row_idx

        # 确定当前系统行对应的子目录（按系统名匹配）
        cur_sys_dir = None
        if is_multi:
            cur_sys = (all_projects[row_idx].system_name or "").replace("/", "_").replace("\\", "_")
            for sd in sys_subdirs:
                sd_name = _os.path.basename(sd)
                if cur_sys and (cur_sys in sd_name or sd_name in cur_sys):
                    cur_sys_dir = sd
                    break

        # 每行独立搜索: 报告打印目录 > 本系统子目录 > 根目录 > 其他子目录
        row_dirs = [report_dir]
        if cur_sys_dir:
            row_dirs.append(cur_sys_dir)
        row_dirs.append(root)
        for sd in sys_subdirs:
            if sd not in row_dirs:
                row_dirs.append(sd)

        for col, keywords in link_map.items():
            found = False
            for kw in keywords:
                for d in row_dirs:
                    if not _os.path.isdir(d):
                        continue
                    for fname in _os.listdir(d):
                        if kw in fname:
                            if col in ext_filter and not ext_filter[col](fname):
                                continue  # 跳过不符合扩展名要求的文件
                            fpath = _os.path.join(d, fname)
                            cell = ws[f'{col}{row_num}']
                            cell.value = fname
                            cell.font = Font(color='0563C1', underline='single', size=10)
                            ole_entries.append((col, row_num, fpath))
                            found = True
                            break
                    if found:
                        break
                if found:
                    break

    # 返回未保存的 workbook，由 _create_report_xlsx_data 统一保存
    return wb, ws, ole_entries
