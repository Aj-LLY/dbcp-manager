"""
项目卡片模块 - 在看板列中以卡片形式展示单个等保测评项目

本模块是系统的核心 UI 组件之一，包含以下主要部分：

  1. _ToolTip 类 - 简易鼠标悬浮提示组件（tooltip），用于文件操作按钮的提示
  2. _add_tooltip 辅助函数 - 快捷创建工具提示
  3. _show_report_dialog 函数 - 报告打印前的编辑确认对话框（14 个可编辑字段）
  4. ProjectCard 类 - 项目卡片主组件，在看板列中展示单个项目的摘要信息

ProjectCard 核心功能：
  - 项目信息展示：公司名称、系统名称、等级、属地、证书编号、截止日期
  - 左侧颜色条：根据截止日期显示绿色（正常）/ 橙色（临近）/ 红色（超期）/ 灰色（无日期）
  - 阶段移动：左箭头（◀）/ 右箭头（▶）按钮，快速将项目移到上/下一阶段
  - 操作按钮：详情、编辑、复制、打开文件夹、重命名、打包、报告打印
  - 选中高亮：单击卡片显示蓝色加粗边框（全局单选）
  - 悬停效果：鼠标进入时背景色变化提示可交互

技术要点：
  - 事件冒泡：通过递归绑定将单击/双击/进入/离开事件绑定到所有子组件（按钮除外）
  - 递归背景色设置：选中/悬停时批量更新子组件背景色（排除固定颜色的功能按钮）
  - 报告打印：弹出编辑确认对话框 -> 创建 XLSX 打印信息表 -> 复制相关文件

依赖关系：
  - models.Project：项目数据实体类
  - utils.Config：全局 UI 配置（字体、颜色、尺寸等）
  - utils.helpers.bordered_entry：边框输入框工厂函数（用于报告对话框）
  - openpyxl：Excel 文件读写（用于生成报告打印信息表）
"""

# =============================================================================
# 导入区
# =============================================================================

import tkinter as tk  # Python 标准 GUI 库，提供 Frame、Label、Button、Canvas 等组件
from datetime import date  # 日期类，用于截止日期的计算和比较

# ---- 模型层 ----
from models.project import Project  # 项目实体类：包含名称、编号、截止日期、阶段归属等所有字段

# ---- 工具层 ----
from utils.config import Config  # 全局配置类：提供字体、颜色、尺寸预警天数等 UI 常量


# =============================================================================
# _ToolTip 类 - 简易按钮悬浮提示组件
# =============================================================================

class _ToolTip:
    """简易按钮悬浮提示 - 鼠标悬停时在按钮上方显示文字提示气泡

    工作原理：
      1. 绑定 <Enter> 事件：创建无边框 Toplevel 窗口，显示提示文字
      2. 绑定 <Leave> 事件：销毁 Toplevel 窗口，移除提示

    使用示例：
      _ToolTip(my_button, "这是一条提示信息")

    Attributes:
        widget (tk.Widget): 被绑定提示的 Tkinter 组件
        text (str): 要显示的提示文字
        tw (tk.Toplevel | None): 提示气泡窗口实例（显示时创建，离开时销毁）
    """

    def __init__(self, widget, text):
        """初始化工具提示并自动绑定事件

        在传入的 widget 上绑定鼠标进入和离开事件，
        用户无需手动管理事件绑定。

        Args:
            widget: 需要显示提示的 Tkinter 组件（如 Button、Label 等）
            text: 鼠标悬停时显示的提示文字
        """
        self.widget = widget  # 保存目标组件引用
        self.text = text  # 保存提示文字
        self.tw = None  # 提示气泡 Toplevel 窗口引用（开始时为 None）
        widget.bind("<Enter>", self._enter)  # 鼠标进入目标组件 -> 显示提示
        widget.bind("<Leave>", self._leave)  # 鼠标离开目标组件 -> 隐藏提示

    def _enter(self, event=None):
        """鼠标进入事件：创建并显示提示气泡窗口

        气泡窗口定位在目标组件的下方约 22px 处，右偏移 5px，
        使用深色背景（#333）+ 白色文字的样式。

        Args:
            event: Tkinter 的 Enter 事件对象（可选）
        """
        # 计算气泡显示位置（相对于屏幕坐标）
        x, y = self.widget.winfo_rootx() + 5, self.widget.winfo_rooty() + 22  # 目标组件左上角 + 偏移
        self.tw = tk.Toplevel(self.widget)  # 创建新的顶级窗口（独立于主窗口）
        self.tw.wm_overrideredirect(True)  # 去除窗口装饰（标题栏、边框），仅显示内容
        self.tw.wm_geometry(f"+{x}+{y}")  # 设置气泡窗口的屏幕位置
        label = tk.Label(self.tw, text=self.text, bg="#333", fg="white",  # 深灰背景，白色文字
                         font=("Microsoft YaHei", 8), padx=4, pady=1)  # 小号字体，紧凑内边距
        label.pack()  # 打包标签到气泡窗口

    def _leave(self, event=None):
        """鼠标离开事件：销毁提示气泡窗口

        Args:
            event: Tkinter 的 Leave 事件对象（可选）
        """
        if self.tw:  # 如果气泡窗口存在
            self.tw.destroy()  # 销毁窗口并释放资源
            self.tw = None  # 清空引用，避免重复销毁


# =============================================================================
# _add_tooltip 辅助函数 - 快捷创建工具提示
# =============================================================================

def _add_tooltip(widget, text):
    """为指定组件快速添加鼠标悬浮提示

    这是 _ToolTip 类的便捷封装函数，一行代码即可完成提示绑定。

    Args:
        widget: 需要添加提示的 Tkinter 组件
        text: 悬浮时显示的提示文字
    """
    _ToolTip(widget, text)  # 创建 _ToolTip 实例（自动绑定事件）


# =============================================================================
# _show_report_dialog 函数 - 报告打印编辑确认对话框
# =============================================================================

def _show_report_dialog(parent, cname="", sname="", location="", deadline=""):
    """显示报告打印信息的编辑确认对话框

    此对话框用于在打印测评报告前，让用户确认和编辑以下 14 个字段：
      1. 客户公司全称
      2. 合同编号或项目名称
      3. 所属地
      4. 系统名称
      5. 是否录入 CRM
      6. 编制/审核/批准日期
      7. 编制人
      8. 审核人
      9. 渗透人员
      10. 测评结论及重大风险隐患数量
      11. 盖章
      12. 打印要求
      13. 项目组长联系人
      14. 实际报告编制人

    特色功能：
      - 支持设置默认值（保存到 data/report_defaults.json）
      - 空值自动回退到项目已有数据兜底
      - 可滚动内容区域（Canvas + Scrollbar）
      - 模态窗口（grab_set() 禁止操作父窗口）

    Args:
        parent: 父级窗口（用于定位和模态绑定）
        cname: 预填的客户公司名称（来自 project.company_name）
        sname: 预填的系统名称（来自 project.system_name）
        location: 预填的所属地（来自 project.location，取省/市前缀）
        deadline: 预填的日期（来自 project.deadline，或当天日期）

    Returns:
        dict | None: 用户确认后返回 14 个字段的字典，取消返回 None
    """
    from tkinter import messagebox  # 消息弹窗，用于保存默认值后的提示

    # ---- 创建对话框主窗口 ----
    dlg = tk.Toplevel(parent)  # 创建子顶级窗口
    dlg.title("报告打印信息确认")  # 设置窗口标题
    dlg.geometry("520x680")  # 设置初始窗口尺寸
    dlg.minsize(440, 500)  # 设置最小尺寸限制
    dlg.configure(bg="#ffffff")  # 白色背景
    dlg.grab_set()  # 设置为模态窗口（禁止操作父窗口）
    dlg.resizable(True, True)  # 允许用户自由调整窗口大小

    # 用于存储用户确认结果的闭包变量
    result = {"confirmed": False}  # 初始为未确认状态

    # ---- 加载保存的默认值 ----
    import json, os  # JSON 解析库和文件系统操作库
    from utils.config import Config  # 全局配置（获取数据目录）
    from utils.helpers import bordered_entry  # 带边框的输入框工厂函数

    defaults = {}  # 存储从文件加载的默认值
    def_path = os.path.join(Config.get_data_dir(), "data", "report_defaults.json")  # 默认值文件路径
    if os.path.exists(def_path):  # 如果文件存在
        try:
            with open(def_path, "r", encoding="utf-8") as f:  # 以 UTF-8 编码打开
                defaults = json.load(f)  # 加载 JSON 内容到字典
        except Exception:  # 任何解析异常（JSON 格式错误、权限不足等）
            pass  # 静默处理：保持空字典，不影响正常使用

    # ========== 底部按钮区域（先 pack 以确保障空间） ==========
    btn_frame = tk.Frame(dlg, bg="#f0f2f5")  # 按钮栏容器 Frame，浅灰背景
    btn_frame.pack(fill=tk.X, side=tk.BOTTOM)  # 固定在对话框底部，水平填充
    tk.Frame(btn_frame, bg="#d0d5dd", height=1).pack(fill=tk.X)  # 1px 灰色分割线
    btn_inner = tk.Frame(btn_frame, bg="#f0f2f5")  # 按钮内容区域（提供内边距）
    btn_inner.pack(fill=tk.X, padx=16, pady=8)  # 水平填充，16px 水平 + 8px 垂直边距

    # ========== 可滚动内容区域 ==========
    canvas = tk.Canvas(dlg, bg="#ffffff", highlightthickness=0)  # Canvas 滚动容器，白色背景，无高亮
    scrollbar = tk.Scrollbar(dlg, orient=tk.VERTICAL, command=canvas.yview)  # 垂直滚动条
    canvas.configure(yscrollcommand=scrollbar.set)  # 绑定 Canvas 与滚动条双向联动
    scrollbar.pack(side=tk.RIGHT, fill=tk.Y)  # 滚动条固定在右侧，垂直填充
    canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)  # Canvas 填充左侧剩余空间

    main = tk.Frame(canvas, bg="#ffffff", padx=20, pady=15)  # 内容主容器，白色背景，20px 水平 + 15px 垂直内边距
    cw = canvas.create_window((0, 0), window=main, anchor="nw")  # 在 Canvas 中创建窗口（嵌入 main Frame）
    main.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))  # 内容变化时更新滚动区域
    canvas.bind("<Configure>", lambda e: canvas.itemconfig(cw, width=e.width))  # Canvas 宽度变化时同步窗口宽度
    # 鼠标滚轮绑定（Canvas 和对话框都支持滚轮滚动）
    canvas.bind("<MouseWheel>", lambda e: canvas.yview_scroll(int(-e.delta/120), "units"))
    dlg.bind("<MouseWheel>", lambda e: canvas.yview_scroll(int(-e.delta/120), "units"))

    # ========== 创建表单字段的工具函数 ==========
    def _make_row(label, default=""):
        """创建一行表单：标签 + 带边框的输入框

        Args:
            label: 字段标签文字（如"客户公司全称"）
            default: 输入框的默认值

        Returns:
            tk.StringVar: 输入框绑定的文本变量（用于读取用户输入）
        """
        tk.Label(main, text=label, bg="#ffffff",  # 字段标签，白色背景
                 font=(Config.FONT_FAMILY, Config.FONT_SIZE_NORMAL)).pack(anchor="w")  # 左对齐
        var = tk.StringVar(value=default)  # 创建文本变量并设置默认值
        _, outer = bordered_entry(main, textvariable=var)  # 创建带边框的输入框
        outer.pack(fill=tk.X, pady=(2, 6))  # 水平填充，上下边距
        return var  # 返回文本变量供后续读取

    # ========== 对话框标题 ==========
    tk.Label(main, text="报告打印信息确认", bg="#ffffff", fg="#2c3e50",  # 深色标题文字
             font=(Config.FONT_FAMILY, Config.FONT_SIZE_HEADER, "bold"),  # 大号粗体
             ).pack(anchor="w", pady=(0, 10))  # 左对齐，下方 10px 间距

    # ========== 创建 14 个字段输入行 ==========
    # 优先使用已保存的默认值，否则使用项目数据；两者都无则留空
    v_cname = _make_row("客户公司全称", defaults.get("cname") or cname)
    v_contract = _make_row("合同编号或项目名称", defaults.get("contract") or f"{cname}网络安全等级保护测评服务项目")
    v_location = _make_row("所属地", defaults.get("location") or location)
    v_sname = _make_row("系统名称", defaults.get("sname") or sname)
    v_crm = _make_row("是否录入CRM", defaults.get("crm") or "是")  # 默认"是"
    v_deadline = _make_row("编制/审核/批准日期", defaults.get("deadline") or deadline)
    v_author = _make_row("编制人", defaults.get("author") or "")
    v_reviewer = _make_row("审核人", defaults.get("reviewer") or "")
    v_pentester = _make_row("渗透人员", defaults.get("pentester") or "")
    v_conclusion = _make_row("测评结论及重大风险隐患数量", defaults.get("conclusion") or "")
    v_seal = _make_row("盖章", defaults.get("seal") or "")
    v_print_req = _make_row("打印要求", defaults.get("print_req") or "")
    v_leader = _make_row("项目组长联系人", defaults.get("leader") or "")
    v_actual = _make_row("实际报告编制人", defaults.get("actual_author") or "")

    # ========== 辅助函数：收集所有字段的值 ==========
    def _collect():
        """收集 14 个字段的当前值并返回字典

        Returns:
            dict: 包含所有字段名称-值映射的字典
        """
        return {
            "cname": v_cname.get().strip(),  # 去除首尾空白
            "contract": v_contract.get().strip(),
            "location": v_location.get().strip(),
            "sname": v_sname.get().strip(),
            "crm": v_crm.get().strip(),
            "deadline": v_deadline.get().strip(),
            "author": v_author.get().strip(),
            "reviewer": v_reviewer.get().strip(),
            "pentester": v_pentester.get().strip(),
            "conclusion": v_conclusion.get().strip(),
            "seal": v_seal.get().strip(),
            "print_req": v_print_req.get().strip(),
            "leader": v_leader.get().strip(),
            "actual_author": v_actual.get().strip(),
        }

    # ========== 确认按钮的回调函数 ==========
    def _on_confirm():
        """用户点击"确认"按钮的处理

        收集所有字段值，对空字段进行兜底填充（使用项目数据），
        设置 confirmed = True 并关闭对话框。
        """
        data = _collect()  # 收集所有字段的当前值

        # --- 空值回退机制：优先用户输入，否则用项目数据兜底 ---
        if not data["cname"]: data["cname"] = cname  # 公司名称回退
        if not data["location"]: data["location"] = location  # 所属地回退
        if not data["sname"]: data["sname"] = sname  # 系统名称回退
        if not data["deadline"]: data["deadline"] = deadline  # 日期回退
        if not data["contract"]: data["contract"] = f"{cname}网络安全等级保护测评服务项目"  # 合同回退

        result["confirmed"] = True  # 标记为已确认
        result["data"] = data  # 存储收集到的数据
        dlg.destroy()  # 关闭对话框

    # ========== 设置默认值按钮的回调函数 ==========
    def _open_defaults_editor():
        """打开设置默认值的独立编辑窗口

        功能：
          1. 读取已保存的默认值文件（data/report_defaults.json）
          2. 如果无保存值，使用当前对话框中的数据
          3. 允许用户编辑所有 14 个字段的默认值
          4. 保存到 JSON 文件
          5. 回填到主对话框的对应字段
        """
        import json, os  # JSON 解析和文件系统操作
        from utils.config import Config  # 配置类（获取数据目录路径）

        # 加载已有默认值
        path = os.path.join(Config.get_data_dir(), "data", "report_defaults.json")
        saved = {}  # 存储已保存的默认值
        if os.path.exists(path):  # 文件存在
            try:
                with open(path, "r", encoding="utf-8") as f:
                    saved = json.load(f)  # 加载 JSON
            except Exception:
                pass  # 解析失败则保持空字典

        # 如果没有任何保存值，用当前对话框的数据预填
        if not any(saved.values()):
            saved = _collect()

        # ---- 创建设置默认值的编辑窗口 ----
        dedit = tk.Toplevel(dlg)  # 子窗口（父窗口为 dlg）
        dedit.title("设置默认值")
        dedit.geometry("500x600")  # 初始尺寸
        dedit.minsize(420, 400)  # 最小尺寸限制
        dedit.configure(bg="#ffffff")  # 白色背景
        dedit.grab_set()  # 模态窗口

        # --- 底部按钮区域（先 pack 以确保障空间） ---
        dbtn_frame = tk.Frame(dedit, bg="#f0f2f5")  # 按钮容器
        dbtn_frame.pack(fill=tk.X, side=tk.BOTTOM)  # 固定在底部
        tk.Frame(dbtn_frame, bg="#d0d5dd", height=1).pack(fill=tk.X)  # 1px 分割线
        dbtn_inner = tk.Frame(dbtn_frame, bg="#f0f2f5")  # 按钮内容区域
        dbtn_inner.pack(fill=tk.X, padx=16, pady=8)

        # --- 可滚动内容区域 ---
        dcanvas = tk.Canvas(dedit, bg="#ffffff", highlightthickness=0)  # Canvas 容器
        dscroll = tk.Scrollbar(dedit, orient=tk.VERTICAL, command=dcanvas.yview)  # 滚动条
        dcanvas.configure(yscrollcommand=dscroll.set)  # 双向联动
        dscroll.pack(side=tk.RIGHT, fill=tk.Y)
        dcanvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        dmain = tk.Frame(dcanvas, bg="#ffffff", padx=20, pady=15)  # 内容主容器
        dw = dcanvas.create_window((0, 0), window=dmain, anchor="nw")  # 嵌入 Canvas
        dmain.bind("<Configure>", lambda e: dcanvas.configure(scrollregion=dcanvas.bbox("all")))
        dcanvas.bind("<Configure>", lambda e: dcanvas.itemconfig(dw, width=e.width))
        # 滚轮支持
        dcanvas.bind("<MouseWheel>", lambda e: dcanvas.yview_scroll(int(-e.delta/120), "units"))
        dedit.bind("<MouseWheel>", lambda e: dcanvas.yview_scroll(int(-e.delta/120), "units"))

        # --- 创建 14 个字段的输入行 ---
        dvars = {}  # 存储所有字段的 tk.StringVar
        for label, key in [
            ("客户公司全称", "cname"), ("合同编号或项目名称", "contract"),
            ("所属地", "location"), ("系统名称", "sname"), ("是否录入CRM", "crm"),
            ("编制/审核/批准日期", "deadline"), ("编制人", "author"),
            ("审核人", "reviewer"), ("渗透人员", "pentester"),
            ("测评结论及重大风险隐患数量", "conclusion"), ("盖章", "seal"),
            ("打印要求", "print_req"), ("项目组长联系人", "leader"),
            ("实际报告编制人", "actual_author"),
        ]:
            tk.Label(dmain, text=label, bg="#ffffff",
                     font=(Config.FONT_FAMILY, Config.FONT_SIZE_NORMAL)).pack(anchor="w")  # 字段标签
            var = tk.StringVar(value=saved.get(key) or "")  # 文本变量，填充保存值或空
            from utils.helpers import bordered_entry  # 边框输入框工厂
            _, outer = bordered_entry(dmain, textvariable=var, width=40)  # 创建输入框，宽 40 字符
            outer.pack(fill=tk.X, pady=(2, 5))  # 水平填充
            dvars[key] = var  # 存储变量引用

        def _save_and_close():
            """保存默认值并关闭编辑窗口"""
            data = {k: v.get().strip() for k, v in dvars.items()}  # 收集所有字段值
            os.makedirs(os.path.dirname(path), exist_ok=True)  # 确保目录存在
            with open(path, "w", encoding="utf-8") as f:  # 写入 JSON 文件
                json.dump(data, f, ensure_ascii=False, indent=2)  # 保留中文，2 空格缩进

            # 回填到主对话框的对应字段
            for key, var in dvars.items():
                target = {
                    "cname": v_cname, "contract": v_contract, "location": v_location,
                    "sname": v_sname, "crm": v_crm, "deadline": v_deadline,
                    "author": v_author, "reviewer": v_reviewer, "pentester": v_pentester,
                    "conclusion": v_conclusion, "seal": v_seal,
                    "print_req": v_print_req, "leader": v_leader,
                    "actual_author": v_actual,
                }.get(key)  # 根据 key 获取主对话框的对应字段变量
                if target:  # 找到对应字段
                    target.set(data[key])  # 将保存的值设置到主对话框

            messagebox.showinfo("提示", "默认值已保存", parent=dedit)  # 提示保存成功
            dedit.destroy()  # 关闭编辑窗口

        # --- 编辑窗口的按钮 ---
        tk.Button(dbtn_inner, text="取消", command=dedit.destroy,
                  bg="#ffffff", fg="#2c3e50", cursor="hand2",
                  font=(Config.FONT_FAMILY, Config.FONT_SIZE_SMALL),
                  relief="flat", padx=20, pady=5,
                  highlightbackground="#d0d5dd", highlightthickness=1,
                  ).pack(side=tk.RIGHT, padx=(10, 0))  # "取消"按钮，右侧
        tk.Button(dbtn_inner, text="保存", command=_save_and_close,
                  bg="#3498db", fg="white", cursor="hand2",
                  font=(Config.FONT_FAMILY, Config.FONT_SIZE_NORMAL, "bold"),
                  relief="flat", padx=20, pady=5,
                  ).pack(side=tk.RIGHT)  # "保存"按钮，右侧
        # 快捷键：Enter 保存，Escape 取消
        dedit.bind("<Return>", lambda e: _save_and_close())
        dedit.bind("<Escape>", lambda e: dedit.destroy())

    # ========== 主对话框的按钮 ==========
    # "设置默认值"按钮（左侧）
    tk.Button(btn_inner, text="设置默认值", command=_open_defaults_editor,
              bg="#f0f2f5", fg="#2c3e50", cursor="hand2",
              font=(Config.FONT_FAMILY, Config.FONT_SIZE_SMALL),
              relief="flat", padx=12, pady=5,
              activebackground="#d5dbdb",  # 按下时的背景色
              ).pack(side=tk.LEFT)

    # "取消"按钮（右侧）
    tk.Button(btn_inner, text="取消", command=dlg.destroy,  # 直接关闭对话框
              bg="#ffffff", fg="#2c3e50", cursor="hand2",
              font=(Config.FONT_FAMILY, Config.FONT_SIZE_SMALL),
              relief="flat", padx=20, pady=5,
              highlightbackground="#d0d5dd", highlightthickness=1,  # 灰色边框
              activebackground="#f0f2f5",
              ).pack(side=tk.RIGHT, padx=(10, 0))

    # "确认"按钮（右侧）
    tk.Button(btn_inner, text="确认", command=_on_confirm,
              bg="#3498db", fg="white", cursor="hand2",  # 蓝色主题
              font=(Config.FONT_FAMILY, Config.FONT_SIZE_NORMAL, "bold"),  # 粗体
              relief="flat", padx=20, pady=5,
              activebackground="#2980b9",  # 按下时的背景色（深蓝）
              ).pack(side=tk.RIGHT)

    # ========== 快捷键绑定 ==========
    dlg.bind("<Return>", lambda e: _on_confirm())  # Enter 键 -> 确认
    dlg.bind("<Escape>", lambda e: dlg.destroy())  # Escape 键 -> 取消（关闭）

    # ========== 等待对话框关闭 ==========
    parent.wait_window(dlg)  # 阻塞等待对话框关闭（模态行为）

    # 返回结果
    if result["confirmed"]:  # 用户点击了确认
        return result["data"]  # 返回 14 个字段的字典
    return None  # 用户取消，返回 None


# =============================================================================
# ProjectCard 类 - 项目卡片主组件
# =============================================================================

class ProjectCard(tk.Frame):
    """项目卡片组件 - 继承自 tk.Frame，作为看板列中的项目展示卡片

    卡片布局（从左到右）：
      ┌──┬────┬──────────────────────────────┬────┬──┐
      │色│ ◀  │         内容区域              │ ▶  │  │
      │条│    │  公司名称 / 系统名称           │    │  │
      │  │    │  等级 · 属地 · 证书编号        │    │  │
      │4 │    │  截止日期                     │    │  │
      │p │    │  ┌──────┬──────┬──────┐      │    │  │
      │x │    │  │ 详情 │ 编辑 │ 复制 │      │    │  │
      │  │    │  └──────┴──────┴──────┘      │    │  │
      │  │    │  ┌───┬───┬───┬───┐         │    │  │
      │  │    │  │📂│📝│📦│📄│         │    │  │
      │  │    │  └───┴───┴───┴───┘         │    │  │
      └──┴────┴──────────────────────────────┴────┴──┘

    交互行为：
      - 单击卡片                -> 选中（蓝色加粗边框）
      - 双击卡片                -> 打开编辑对话框
      - 点击 ◀ 箭头             -> 项目移至上一阶段
      - 点击 ▶ 箭头             -> 项目移至下一阶段
      - 点击"详情"按钮          -> 打开项目详情窗口
      - 点击"编辑"按钮          -> 打开编辑对话框（同双击）
      - 点击"复制"按钮          -> 创建项目副本
      - 点击 📂 按钮            -> 打开项目文件夹
      - 点击 📝 按钮            -> 一键重命名项目文件
      - 点击 📦 按钮            -> 打包过程文档为 ZIP
      - 点击 📄 按钮            -> 报告打印（生成 XLSX）

    Attributes:
        project (Project): 关联的项目实体对象
        is_selected (bool): 当前是否处于选中状态
        on_click: 单击回调函数 -> (card)
        on_double_click: 双击回调函数 -> (card)
        on_detail: 详情按钮回调 -> (card)
        on_edit: 编辑按钮回调 -> (card)
        on_copy: 复制按钮回调 -> (card)
        on_move_prev: 左箭头回调 -> (card)
        on_move_next: 右箭头回调 -> (card)
    """

    def __init__(self, parent, project: Project, **kwargs):
        """初始化项目卡片组件

        Args:
            parent: 父级容器（通常为 KanbanColumn 中的 cards_frame）
            project: 关联的项目实体对象（包含所有业务字段）
            **kwargs: 传递给父类 tk.Frame 的额外关键字参数
        """
        # 调用父类构造方法，设置卡片默认样式
        super().__init__(parent, bg=Config.CARD_BG,  # 白色卡片背景
                         highlightbackground=Config.CARD_BORDER,  # 浅灰边框
                         highlightthickness=1,  # 1px 边框宽度
                         cursor="hand2",  # 手型光标提示可点击
                         **kwargs)

        # ---- 公共属性 ----
        self.project = project  # 保存关联的项目实体引用（只读，外部不应直接修改）
        self.is_selected = False  # 选中状态标志（初始为未选中）

        # ---- 回调函数指针（由 KanbanBoard 创建卡片后设置） ----
        self.on_click = None  # 单击 -> 选中/取消选中
        self.on_double_click = None  # 双击 -> 打开编辑对话框
        self.on_detail = None  # "详情"按钮 -> 打开详情窗口
        self.on_edit = None  # "编辑"按钮 -> 打开编辑对话框
        self.on_copy = None  # "复制"按钮 -> 创建项目副本
        self.on_move_prev = None  # ◀ 左箭头 -> 移至上一阶段
        self.on_move_next = None  # ▶ 右箭头 -> 移至下一阶段

        # ---- 构建 UI 和绑定事件 ----
        self._build_ui()  # 构建卡片内部的完整 UI 布局
        self._bind_events()  # 为所有子组件递归绑定鼠标事件

    # ==================================================================================
    # UI 构建
    # ==================================================================================

    def _build_ui(self):
        """构建卡片完整布局

        布局顺序（从左到右）：
          1. 左侧颜色条（4px，状态指示）
          2. 左箭头按钮（◀）
          3. 中间内容区域（项目信息 + 操作按钮）
          4. 右箭头按钮（▶）
        """
        # 根据截止日期获取状态指示颜色
        status_color = self._get_status_color()  # 绿色=正常 / 橙色=临近 / 红色=超期 / 灰色=无日期

        # ========== 1. 左侧颜色条（4px 宽垂直色条） ==========
        self._color_bar = tk.Frame(self, bg=status_color, width=4)  # 状态色背景，固定 4px 宽
        self._color_bar.pack(side=tk.LEFT, fill=tk.Y)  # 左侧，垂直填充
        self._color_bar.pack_propagate(False)  # 禁止内容撑开，保持 4px 宽度

        # ========== 2. 左箭头按钮（◀） ==========
        self._prev_btn = tk.Button(
            self, text="\u25c0",  # Unicode ◀ 字符
            command=self._on_prev_click,  # 点击 -> 调用左箭头处理
            bg=Config.CARD_BG, fg="#b0b8c1",  # 卡片背景色，浅灰文字
            font=(Config.FONT_FAMILY, 8),  # 小号字体
            relief="flat", borderwidth=0,  # 扁平样式，无边框
            cursor="hand2", padx=3, pady=0,  # 手型光标，紧凑内边距
            activebackground=Config.CARD_HOVER_BG, activeforeground="#2c3e50",  # 悬停时高亮
        )
        self._prev_btn.pack(side=tk.LEFT, fill=tk.Y)  # 左侧，垂直填充

        # ========== 3. 中间内容区域 ==========
        self._content = tk.Frame(self, bg=Config.CARD_BG)  # 内容容器，白色背景
        self._content.pack(side=tk.LEFT, fill=tk.BOTH, expand=True,  # 填充剩余空间
                           padx=6, pady=(8, 4))  # 水平 6px 内边距，上方 8px 下方 4px

        # --- 3a. 系统名称（居中、粗体主标题） ---
        sys_display = self.project.system_name or self.project.company_name or "\u65e0\u540d\u79f0"  # \u65e0\u540d\u79f0 = 无名称
        if len(sys_display) > 12:  # 名称过长截断
            sys_display = sys_display[:11] + "\u2026"  # \u2026 = …（省略号）
        self._sys_label = tk.Label(
            self._content, text=sys_display, bg=Config.CARD_BG,  # 卡片背景色
            font=(Config.FONT_FAMILY, Config.FONT_SIZE_NORMAL, "bold"),  # 正常字号加粗
            anchor="center", fg="#2c3e50",  # 文字居中，深灰色
        )
        self._sys_label.pack(fill=tk.X)  # 水平填充

        # --- 3b. 公司名称（仅当两者都有时作为副标题显示） ---
        if self.project.system_name and self.project.company_name:
            company_display = self.project.company_name
            if len(company_display) > 14:  # 公司名称过长截断
                company_display = company_display[:13] + "\u2026"  # … 省略号
            self._company_label = tk.Label(
                self._content, text=company_display, bg=Config.CARD_BG,
                font=(Config.FONT_FAMILY, Config.FONT_SIZE_SMALL),  # 小字号
                anchor="center", fg="#5d6d7e",  # 居中，中灰色
            )
            self._company_label.pack(fill=tk.X)
        else:
            self._company_label = None  # 无公司名称时置空

        # --- 3c. 系统安全等级（居中、小字、紫色） ---
        if self.project.level:
            self._level_label = tk.Label(
                self._content, text=self.project.level,  # 显示等级（如"三级"）
                bg=Config.CARD_BG,
                font=(Config.FONT_FAMILY, Config.FONT_SIZE_SMALL - 1),  # 较小字号
                anchor="center", fg="#8e44ad",  # 紫色文字
            )
            self._level_label.pack(fill=tk.X)
        else:
            self._level_label = None  # 无等级时不显示

        # --- 3d. 所属地（居中、小字、灰色） ---
        if self.project.location:
            self._loc_label = tk.Label(
                self._content, text=self.project.location,  # 显示所在地
                bg=Config.CARD_BG,
                font=(Config.FONT_FAMILY, Config.FONT_SIZE_SMALL - 1),
                anchor="center", fg="#7f8c8d",  # 灰蓝文字
            )
            self._loc_label.pack(fill=tk.X)
        else:
            self._loc_label = None  # 无属地时不显示

        # --- 3e. 证书编号 / 备案状态（居中） ---
        if self.project.cert_number:  # 有证书编号 -> 已备案状态
            cert_display = self.project.cert_number
            if len(cert_display) > 18:  # 编号过长截断
                cert_display = cert_display[:17] + "\u2026"
            self._cert_label = tk.Label(
                self._content, text="\U0001f4dc \u5df2\u5907\u6848 " + cert_display,  # 📜 已备案 + 编号
                bg=Config.CARD_BG,
                font=(Config.FONT_FAMILY, Config.FONT_SIZE_SMALL - 1),
                anchor="center", fg="#27ae60",  # 绿色
            )
        else:  # 无证书编号 -> 未备案状态
            self._cert_label = tk.Label(
                self._content, text="\U0001f4dc \u672a\u5907\u6848",  # 📜 未备案
                bg=Config.CARD_BG,
                font=(Config.FONT_FAMILY, Config.FONT_SIZE_SMALL - 1),
                anchor="center", fg="#e67e22",  # 橙色
            )
        self._cert_label.pack(fill=tk.X)

        # --- 3f. 交付截止日期（居中、带剩余天数提示） ---
        deadline_text = self._format_deadline()  # 格式化的日期文本（含图标和天数）
        fg_color = self._get_deadline_color()  # 根据紧急性获取文字颜色
        self._deadline_label = tk.Label(
            self._content, text=deadline_text, bg=Config.CARD_BG,
            font=(Config.FONT_FAMILY, Config.FONT_SIZE_SMALL),
            anchor="center", fg=fg_color,  # 动态颜色
        )
        self._deadline_label.pack(fill=tk.X, pady=(2, 0))  # 上方 2px 间距

        # --- 3g. 操作按钮第 1 行：详情 / 编辑 / 复制（居中布局） ---
        btn_frame = tk.Frame(self._content, bg=Config.CARD_BG)  # 按钮行容器
        btn_frame.pack(fill=tk.X, pady=(4, 0))  # 水平填充，上方 4px 间距

        # 使用两个弹性空白 Frame 实现居中效果
        tk.Frame(btn_frame, bg=Config.CARD_BG).pack(side=tk.LEFT, expand=True)  # 左侧弹性空白

        # "详情"按钮
        self._detail_btn = tk.Button(
            btn_frame, text="\u8be6\u60c5",  # \u8be6\u60c5 = 详情
            command=self._on_detail_click,
            bg="#ecf0f1", fg="#2c3e50",  # 浅灰背景，深灰文字
            font=(Config.FONT_FAMILY, Config.FONT_SIZE_SMALL),
            relief="flat", borderwidth=0, cursor="hand2", padx=8, pady=1,
            activebackground="#d5dbdb",  # 悬停时更深灰
        )
        self._detail_btn.pack(side=tk.LEFT, padx=(0, 4))  # 右间距 4px

        # "编辑"按钮
        self._edit_btn = tk.Button(
            btn_frame, text="\u7f16\u8f91",  # \u7f16\u8f91 = 编辑
            command=self._on_edit_click,
            bg="#3498db", fg="white",  # 蓝色背景，白色文字
            font=(Config.FONT_FAMILY, Config.FONT_SIZE_SMALL),
            relief="flat", borderwidth=0, cursor="hand2", padx=8, pady=1,
            activebackground="#2980b9",  # 悬停时深蓝
        )
        self._edit_btn.pack(side=tk.LEFT)

        # "复制"按钮
        self._copy_btn = tk.Button(
            btn_frame, text="\u590d\u5236",  # \u590d\u5236 = 复制
            command=self._on_copy_click,
            bg="#27ae60", fg="white",  # 绿色背景，白色文字
            font=(Config.FONT_FAMILY, Config.FONT_SIZE_SMALL),
            relief="flat", borderwidth=0, cursor="hand2", padx=8, pady=1,
            activebackground="#1e8449",  # 悬停时深绿
        )
        self._copy_btn.pack(side=tk.LEFT, padx=(4, 0))  # 左间距 4px

        tk.Frame(btn_frame, bg=Config.CARD_BG).pack(side=tk.LEFT, expand=True)  # 右侧弹性空白

        # --- 3h. 操作按钮第 2 行：文件操作按钮（居中布局） ---
        file_btn_frame = tk.Frame(self._content, bg=Config.CARD_BG)
        file_btn_frame.pack(fill=tk.X, pady=(2, 0))  # 水平填充，上方 2px 间距

        tk.Frame(file_btn_frame, bg=Config.CARD_BG).pack(side=tk.LEFT, expand=True)  # 左侧弹性空白

        # "打开文件夹"按钮（📂）
        self._folder_btn = tk.Button(
            file_btn_frame, text="\U0001f4c2",  # 📂 文件夹图标
            command=self._on_folder_click,
            bg="#ecf0f1", fg="#2c3e50",  # 浅灰背景
            font=(Config.FONT_FAMILY, Config.FONT_SIZE_SMALL),
            relief="flat", borderwidth=0, cursor="hand2", padx=7, pady=1,
            activebackground="#d5dbdb",
        )
        self._folder_btn.pack(side=tk.LEFT, padx=(0, 2))
        _add_tooltip(self._folder_btn, "打开项目文件夹")

        # "项目初始化"按钮（🔧）
        self._init_btn = tk.Button(
            file_btn_frame, text="\U0001f527",  # 🔧
            command=self._on_init_click,
            bg="#f0f2f5", fg="#2c3e50",
            font=(Config.FONT_FAMILY, Config.FONT_SIZE_SMALL),
            relief="flat", borderwidth=0, cursor="hand2", padx=7, pady=1,
            activebackground="#d5dbdb",
        )
        self._init_btn.pack(side=tk.LEFT, padx=(2, 2))
        _add_tooltip(self._init_btn, "项目初始化（创建子目录和模板文件）")

        # "一键重命名"按钮（📝）
        self._rename_btn = tk.Button(
            file_btn_frame, text="\U0001f4dd",  # 📝 备忘录图标
            command=self._on_rename_click,
            bg="#f0f2f5", fg="#2c3e50",
            font=(Config.FONT_FAMILY, Config.FONT_SIZE_SMALL),
            relief="flat", borderwidth=0, cursor="hand2", padx=7, pady=1,
            activebackground="#d5dbdb",
        )
        self._rename_btn.pack(side=tk.LEFT, padx=(2, 2))
        _add_tooltip(self._rename_btn, "批量重命名文件")  # 悬浮提示

        # "打包过程文档"按钮（📦）
        self._zip_btn = tk.Button(
            file_btn_frame, text="\U0001f4e6",  # 📦 包裹图标
            command=self._on_zip_click,
            bg="#f39c12", fg="white",  # 橙色背景
            font=(Config.FONT_FAMILY, Config.FONT_SIZE_SMALL),
            relief="flat", borderwidth=0, cursor="hand2", padx=7, pady=1,
            activebackground="#e67e22",  # 悬停时深橙
        )
        self._zip_btn.pack(side=tk.LEFT, padx=(2, 2))
        _add_tooltip(self._zip_btn, "打包过程文档")  # 悬浮提示

        # "报告打印"按钮（📄）
        self._report_btn = tk.Button(
            file_btn_frame, text="\U0001f4c4",  # 📄 文档图标
            command=self._on_report_print_click,
            bg="#8e44ad", fg="white",  # 紫色背景
            font=(Config.FONT_FAMILY, Config.FONT_SIZE_SMALL),
            relief="flat", borderwidth=0, cursor="hand2", padx=7, pady=1,
            activebackground="#7d3c98",  # 悬停时深紫
        )
        self._report_btn.pack(side=tk.LEFT, padx=(2, 0))
        _add_tooltip(self._report_btn, "报告打印")  # 悬浮提示

        tk.Frame(file_btn_frame, bg=Config.CARD_BG).pack(side=tk.LEFT, expand=True)  # 右侧弹性空白

        # ========== 4. 右箭头按钮（▶） ==========
        self._next_btn = tk.Button(
            self, text="\u25b6",  # Unicode ▶ 字符
            command=self._on_next_click,  # 点击 -> 调用右箭头处理
            bg=Config.CARD_BG, fg="#b0b8c1",
            font=(Config.FONT_FAMILY, 8),
            relief="flat", borderwidth=0,
            cursor="hand2", padx=3, pady=0,
            activebackground=Config.CARD_HOVER_BG, activeforeground="#2c3e50",
        )
        self._next_btn.pack(side=tk.RIGHT, fill=tk.Y)  # 右侧，垂直填充

    # ==================================================================================
    # 日期格式化与状态颜色
    # ==================================================================================

    def _format_deadline(self) -> str:
        """格式化截止日期显示文本

        根据剩余天数生成带图标和提示的日期字符串：
          - 无截止日期：显示"无截止日期"
          - 已超期：日期后显示"(已超期)"
          - 临近截止（<= 7 天）：日期后显示"(剩N天)"
          - 正常：仅显示日期

        Returns:
            str: 格式化后的日期文本，格式如 📅 2026-06-15 (剩5天)
        """
        if not self.project.deadline:  # 项目没有设置截止日期
            return "\u65e0\u622a\u6b62\u65e5\u671f"  # \u65e0\u622a\u6b62\u65e5\u671f = 无截止日期
        text = "\U0001f4c5 " + self.project.deadline  # 📅 日历图标 + 日期字符串
        days_left = self._days_until_deadline()  # 计算距离截止日期的天数
        if days_left < 0:  # 已超过截止日期
            text += " (\u5df2\u8d85\u671f)"  # \u5df2\u8d85\u671f = 已超期
        elif days_left <= Config.DEADLINE_WARNING_DAYS:  # 在预警天数范围内（默认 7 天）
            text += f" (\u5269{days_left}\u5929)"  # \u5269 = 剩, \u5929 = 天
        return text  # 返回完整格式化文本

    def _get_status_color(self) -> str:
        """根据项目截止日期获取左侧颜色条的显示颜色

        颜色含义：
          - 灰色 (#95a5a6): 无截止日期，状态不明
          - 红色 (#e74c3c): 已超期，紧急
          - 橙色 (#f39c12): 临近截止日期（<=7 天），需要关注
          - 绿色 (#2ecc71): 时间充裕，正常进行

        Returns:
            str: CSS 颜色代码
        """
        if not self.project.deadline:  # 无截止日期
            return "#95a5a6"  # 灰色
        days_left = self._days_until_deadline()  # 计算剩余天数
        if days_left < 0:  # 已过期
            return "#e74c3c"  # 红色
        elif days_left <= Config.DEADLINE_WARNING_DAYS:  # 临近截止（<= 7 天）
            return "#f39c12"  # 橙色警告
        return "#2ecc71"  # 绿色正常

    def _get_deadline_color(self) -> str:
        """根据截止日期获取日期标签的文字颜色

        Returns:
            str: CSS 颜色代码（绿/橙/红/灰，与状态色条略有不同的色调）
        """
        if not self.project.deadline:
            return "#95a5a6"  # 灰色
        days_left = self._days_until_deadline()
        if days_left < 0:
            return "#e74c3c"  # 红色
        elif days_left <= Config.DEADLINE_WARNING_DAYS:
            return "#e67e22"  # 橙色
        return "#27ae60"  # 绿色

    def _days_until_deadline(self) -> int:
        """计算距离截止日期的剩余天数

        计算公式：deadline - today（正数 = 未来，负数 = 过去，零 = 今天到期）

        Returns:
            int: 剩余天数，负数为已超期，999 表示无法计算（无日期或格式异常）
        """
        if not self.project.deadline:  # 无截止日期
            return 999  # 返回极大值，在排序和判断中视为"远期"
        try:
            dl = date.fromisoformat(self.project.deadline)  # 将 ISO 格式字符串解析为 date 对象
            return (dl - date.today()).days  # 计算日期差（timedelta.days 返回整数天数）
        except (ValueError, TypeError):  # 日期格式不符合 ISO 标准（如 None、"2026/06/15"等）
            return 999  # 格式化异常时返回极大值，避免崩溃

    # ==================================================================================
    # 按钮点击事件处理
    # ==================================================================================

    def _on_prev_click(self):
        """处理左箭头按钮（◀）的点击事件

        触发外部设置的回调函数，由 KanbanBoard 计算上一阶段的 ID，
        最终由 MainWindow 执行项目阶段移动。
        """
        if self.on_move_prev:  # 回调函数已设置（防御性检查）
            self.on_move_prev(self)  # 调用回调并传入自身引用

    def _on_next_click(self):
        """处理右箭头按钮（▶）的点击事件

        触发外部设置的回调函数，由 KanbanBoard 计算下一阶段的 ID，
        最终由 MainWindow 执行项目阶段移动。
        """
        if self.on_move_next:  # 回调函数已设置（防御性检查）
            self.on_move_next(self)  # 调用回调并传入自身引用

    def _on_detail_click(self):
        """处理"详情"按钮的点击事件 - 打开项目详情窗口"""
        if self.on_detail:
            self.on_detail(self)

    def _on_edit_click(self):
        """处理"编辑"按钮的点击事件 - 打开项目编辑对话框"""
        if self.on_edit:
            self.on_edit(self)

    def _on_copy_click(self):
        """处理"复制"按钮的点击事件 - 复制当前项目创建副本"""
        if self.on_copy:
            self.on_copy(self)

    # ==================================================================================
    # 文件操作按钮事件处理
    # ==================================================================================

    def _on_folder_click(self):
        """处理"打开文件夹"按钮点击 - 在系统文件管理器中打开项目目录

        调用各操作系统的默认文件管理器：
          - Windows：os.startfile(path) 直接打开
          - Linux/macOS：subprocess.run(["xdg-open", path])
        """
        import os, subprocess, sys  # 系统操作和子进程模块
        try:
            path = self._find_project_folder()  # 查找项目文件夹路径
            if path and os.path.isdir(path):  # 路径存在且是目录
                if sys.platform == "win32":  # Windows 系统
                    os.startfile(path)  # 使用 Windows 默认方式打开（类似双击文件夹）
                else:  # 非 Windows 系统（Linux/macOS）
                    subprocess.run(["xdg-open", path])  # 调用 xdg-open 打开
        except Exception:  # 打开失败（权限不足、路径不存在等）
            pass  # 静默处理：打开失败不阻塞主流程

    def _find_project_folder(self) -> str:
        """根据项目信息查找本地文件夹路径

        查找策略（按优先级从高到低）：
          1. 项目存储的 folder_path 属性（最直接，优先使用）
          2. 按公司名 + 系统名 + 创建日期的关键词模糊搜索（兜底方案）

        搜索关键词：
          - 公司名称（清理路径非法字符后）
          - 系统名称（清理路径非法字符后）
          - 创建日期（YYMMDD 格式，取项目 created_at 前 10 位的后 6 位）

        Returns:
            str: 找到的文件夹路径，未找到返回空字符串 ""
        """
        import os, re  # 文件系统和正则模块
        from utils.config import Config  # 配置类（获取数据目录）

        # 策略 1：优先使用项目存储的文件夹路径
        if self.project.folder_path and os.path.isdir(self.project.folder_path):
            return self.project.folder_path  # 直接返回存储的路径

        # 策略 2：按关键词搜索
        base = Config.get_data_dir()  # 获取程序数据根目录
        if not os.path.exists(base):  # 根目录不存在
            return ""  # 无数据目录，返回空
        # 清理名称中的路径非法字符，统一替换为下划线
        cname = (self.project.company_name or "未命名").replace("/", "_").replace("\\", "_")
        sname = (self.project.system_name or "").replace("/", "_").replace("\\", "_")
        date_str = ""  # 日期关键词（默认为空，不参与过滤）
        if self.project.created_at:  # 有创建时间
            date_str = self.project.created_at[:10].replace("-", "")[2:]  # YYYY-MM-DD -> YYMMDD

        # 遍历数据目录，查找匹配的文件夹名
        for name in os.listdir(base):
            full = os.path.join(base, name)  # 拼接完整路径
            if not os.path.isdir(full):  # 非目录跳过
                continue
            # 匹配规则：名称中包含公司名，且（无系统名 或 包含系统名），且（无日期 或 包含日期）
            if cname in name and (not sname or sname in name) and (not date_str or date_str in name):
                return full  # 找到匹配，返回路径
        return ""  # 未找到任何匹配

    # ==================================================================================
    # _on_rename_click - 一键重命名功能
    # ==================================================================================

    def _on_rename_click(self):
        """处理"一键重命名"按钮点击 - 批量修正项目过程文件的命名

        此方法是系统的核心自动化功能之一，执行以下流程：

        1. 查找项目文件夹（按 folder_path 或关键词搜索）
        2. 解压 ZIP 文件并重命名内容：
           - "测评方案评审记录表.zip" -> 解压提取文件
           - "测评报告评审表.zip" -> 解压提取文件（保留终审，删除初审）
        3. 删除包含"初审"的文件
        4. 修正文件命名格式为：{编号}-{公司}-{系统}-{标准名称}.{扩展名}
           - 已带编号的文件（如 "07-测评方案.docx"）：匹配关键词后更正编号和名称
           - 无编号的文件：自动添加编号和名称前缀
        5. 修正子目录命名格式：
           - "报告打印" -> "00-{公司}-{系统}-报告打印"
           - "渗透测试报告" -> "13-{公司}-{系统}-渗透测试报告"
        6. 输出操作报告（显示处理了多少文件、跳过多少文件）

        关键字映射表（key_map）：
          {匹配关键词: (编号, 标准化名称)}，按关键词长度倒序匹配避免误匹配
        """
        import os, re, zipfile, shutil  # 操作系统、正则、ZIP 压缩、文件操作模块
        from tkinter import messagebox  # 消息弹窗

        try:
            root = self._find_project_folder()  # 查找项目文件夹路径
            if not root or not os.path.isdir(root):  # 文件夹不存在
                messagebox.showinfo("提示", "未找到项目文件夹")
                return

            # 生成文件名的前导前缀
            cname = (self.project.company_name or "未命名").replace("/", "_").replace("\\", "_")
            sname = (self.project.system_name or "").replace("/", "_").replace("\\", "_")
            new_prefix = f"{cname}-{sname}"  # 格式：公司名-系统名

            # ---- 关键字映射表 ----
            # 格式：{文件名关键词: (编号, 标准化显示名称)}
            # 按关键词长度倒序匹配，防止"测评方案评审表"误匹配为"测评方案"
            key_map = {
                "保密承诺书": ("02", "保密承诺书"),  # 02 号文件
                "测评调研表": ("03", "测评调研表"),  # 03 号文件
                "测评授权书": ("04", "测评授权书"),  # 04 号文件
                "风险告知书": ("05", "风险告知书"),  # 05 号文件
                "项目计划书": ("06", "项目计划书"),  # 06 号文件
                "测评方案": ("07", "测评方案"),  # 07 号文件（注意：需放在"测评方案评审记录表"之后匹配）
                "归档材料评审记录表": ("08", "测评方案评审表"),  # 08 号文件（历史名称）
                "测评方案评审表": ("08", "测评方案评审表"),  # 08 号文件（标准名称）
                "首次会议记录": ("09", "首次会议记录"),  # 09 号文件
                "测评现场记录表": ("10", "测评现场记录表"),  # 10 号文件
                "问题汇总": ("11", "问题汇总及整改建设书"),  # 11 号文件
                "漏洞扫描报告": ("12", "漏洞扫描报告"),  # 12 号文件
                "项目文档移交清单": ("14", "项目文档移交清单"),  # 14 号文件
                "末次会议记录": ("15", "末次会议记录"),  # 15 号文件
                "测评报告-终稿": ("16", "测评报告-终稿"),  # 16 号文件
                "测评报告评审记录表": ("17", "测评报告评审表"),  # 17 号文件（历史名称）
                "测评报告评审表": ("17", "测评报告评审表"),  # 17 号文件（标准名称）
                "服务情况评价表": ("18", "服务情况评价表"),  # 18 号文件
                "报备表": ("19", "报备表"),
                "渗透测试报告": ("13", "渗透测试报告"),  # 渗透测试报告目录
            }

            renamed = 0  # 重命名计数器（累计处理了多少个项目）
            msgs = []  # 操作报告消息列表（每步操作一条记录）

            # ========== 步骤 1: ZIP 文件解压处理 ==========
            for fname in os.listdir(root):  # 遍历项目根目录所有文件
                fpath = os.path.join(root, fname)  # 拼接文件完整路径
                if not os.path.isfile(fpath) or not fname.lower().endswith(".zip"):  # 跳过非 ZIP 文件
                    continue

                # 处理"测评方案评审记录表.zip" -> 解压并重命名内容为 08
                if "测评方案评审记录表" in fname:
                    try:
                        with zipfile.ZipFile(fpath, "r") as zf:  # 以读取模式打开 ZIP
                            zf.extractall(root)  # 解压全部内容到项目根目录
                        os.remove(fpath)  # 删除原 ZIP 文件
                        renamed += 1  # 计数 +1
                        msgs.append(f"解压: {fname} -> 提取文件")  # 记录操作
                    except Exception as e:  # 解压失败（损坏的 ZIP、权限不足等）
                        msgs.append(f"解压失败: {fname} ({e})")

                # 处理"渗透测试报告.zip" -> 解压到独立文件夹后重命名
                if "渗透测试报告" in fname and "评审" not in fname:
                    try:
                        # 解压到临时目录
                        tmp_dir = os.path.join(root, "_渗透测试报告_tmp")
                        if os.path.exists(tmp_dir):
                            shutil.rmtree(tmp_dir)
                        os.makedirs(tmp_dir)
                        with zipfile.ZipFile(fpath, "r") as zf:
                            zf.extractall(tmp_dir)
                        os.remove(fpath)
                        # 去除 ZIP 内的公共顶层目录前缀
                        items = os.listdir(tmp_dir)
                        if len(items) == 1 and os.path.isdir(os.path.join(tmp_dir, items[0])):
                            # ZIP自带一个顶层目录 -> 直接移动该目录
                            src = os.path.join(tmp_dir, items[0])
                            dst = os.path.join(root, f"13-{new_prefix}-渗透测试报告")
                        else:
                            # 文件散落 -> 创建目标目录，移入所有文件
                            dst = os.path.join(root, f"13-{new_prefix}-渗透测试报告")
                            if not os.path.exists(dst):
                                os.makedirs(dst)
                            for item in items:
                                shutil.move(os.path.join(tmp_dir, item), os.path.join(dst, item))
                            shutil.rmtree(tmp_dir)
                            renamed += 1
                            msgs.append(f"解压: {fname} -> 13-{new_prefix}-渗透测试报告/")
                            continue
                        if os.path.exists(dst):
                            shutil.rmtree(dst)
                        shutil.move(src, dst)
                        shutil.rmtree(tmp_dir)
                        renamed += 1
                        msgs.append(f"解压: {fname} -> 13-{new_prefix}-渗透测试报告/")
                    except Exception as e:
                        msgs.append(f"解压失败: {fname} ({e})")

                # 处理"测评报告评审表.zip" -> 解压（后续会删除初审版本）
                elif "测评报告评审表" in fname:
                    try:
                        with zipfile.ZipFile(fpath, "r") as zf:  # 以读取模式打开
                            zf.extractall(root)  # 解压到项目根目录
                        os.remove(fpath)  # 删除原 ZIP
                        renamed += 1
                        msgs.append(f"解压: {fname} -> 提取文件")
                    except Exception as e:
                        msgs.append(f"解压失败: {fname} ({e})")

            # ========== 步骤 2: 删除初审版本文件 ==========
            for fname in os.listdir(root):
                if "初审" in fname:  # 包含"初审"的文件（如测评报告评审表-初审.docx）
                    try:
                        os.remove(os.path.join(root, fname))  # 直接删除
                        msgs.append(f"删除初审: {fname}")
                    except Exception:  # 删除失败（权限等），静默跳过
                        pass

            # ========== 步骤 3: 批量重命名文件 ==========
            for fname in os.listdir(root):  # 遍历根目录所有文件
                fpath = os.path.join(root, fname)
                if not os.path.isfile(fpath) or fname.endswith(".zip"):  # 跳过 ZIP 和目录
                    continue
                name_no_ext, ext = os.path.splitext(fname)  # 分离文件名和扩展名（"07-测评方案", ".docx"）

                # ---- 情况 A: 文件名已有编号前缀（如 "07-测评方案.docx"） ----
                m = re.match(r"^(\d{2})-(.+)", name_no_ext)  # 匹配 "###-内容" 格式
                if m:  # 已有编号前缀
                    num = m.group(1)  # 当前编号（如 "07"）
                    rest = name_no_ext[len(num) + 1:]  # 编号后面的剩余部分（如 "测评方案.docx"）
                    # 按关键词长度从大到小匹配，避免"测评方案评审表"被"测评方案"先匹配
                    for keyword in sorted(key_map, key=len, reverse=True):
                        if keyword in rest:  # 剩余部分中包含关键字
                            num, target_kw = key_map[keyword]  # 获取正确的编号和标准名称
                            new_name = f"{num}-{new_prefix}-{target_kw}{ext}"  # 构建标准文件名
                            if new_name != fname:  # 需要重命名
                                new_path = os.path.join(root, new_name)
                                if not os.path.exists(new_path):  # 目标文件不存在
                                    os.rename(fpath, new_path)  # 执行重命名
                                    renamed += 1
                                else:  # 目标文件已存在（冲突）
                                    msgs.append(f"跳过(已存在): {new_name}")
                            break  # 匹配到一个关键词后退出内层循环

                # ---- 情况 B: 文件名无编号（如 "测评调研表.docx"） ----
                else:
                    for keyword in sorted(key_map, key=len, reverse=True):  # 按关键词长度倒序
                        if keyword in name_no_ext:  # 文件名中包含关键字
                            num, target_kw = key_map[keyword]  # 获取编号和标准名称
                            new_name = f"{num}-{new_prefix}-{target_kw}{ext}"  # 构建标准名
                            new_path = os.path.join(root, new_name)
                            if not os.path.exists(new_path):  # 不冲突
                                os.rename(fpath, new_path)  # 重命名
                                renamed += 1
                            else:
                                msgs.append(f"跳过(已存在): {new_name}")
                            break  # 匹配到一个后退出

            # ========== 步骤 4: 子目录重命名 ==========
            for dname in os.listdir(root):  # 遍历所有目录
                dpath = os.path.join(root, dname)
                if not os.path.isdir(dpath):  # 跳过文件
                    continue
                for keyword, num in {"报告打印": "00", "渗透测试报告": "13"}.items():
                    if keyword in dname and (cname not in dname or sname not in dname):  # 需要更新前缀
                        new_dname = f"{num}-{new_prefix}-{keyword}"  # 标准目录名
                        new_dpath = os.path.join(root, new_dname)
                        if not os.path.exists(new_dpath):  # 不冲突
                            os.rename(dpath, new_dpath)  # 重命名目录
                            renamed += 1
                        break  # 匹配到一个关键词后退出

            # ========== 步骤 5: 显示操作报告 ==========
            if msgs:  # 有详细操作记录
                msg_text = "\n".join(msgs[:15])  # 取前 15 条
                if len(msgs) > 15:  # 超过 15 条则显示统计
                    msg_text += f"\n...共 {len(msgs)} 条"
                messagebox.showinfo("操作报告", msg_text)  # 弹窗展示
            elif renamed:  # 没有详细记录但统计 > 0
                messagebox.showinfo("完成", f"已处理 {renamed} 个项目")
            else:  # 无需任何修改
                messagebox.showinfo("提示", "所有文件名已是最新，无需修改")

        except Exception as e:  # 捕获所有未预见的异常
            messagebox.showerror("错误", f"操作失败: {e}")  # 弹窗显示错误

    # ==================================================================================
    # _on_zip_click - 打包过程文档功能
    # ==================================================================================

    def _on_init_click(self):
        """项目初始化：创建标准子目录和保密承诺书模板"""
        import os
        from tkinter import messagebox
        try:
            root = self._find_project_folder()
            if not root or not os.path.isdir(root):
                messagebox.showinfo("提示", "未找到项目文件夹")
                return
            cname = (self.project.company_name or "未命名").replace("/", "_").replace("\\", "_")
            sname = (self.project.system_name or "").replace("/", "_").replace("\\", "_")
            created = []; existed = []
            # 01-其他归档文件
            dname = "01-其他归档文件"
            dpath = os.path.join(root, dname)
            if not os.path.exists(dpath):
                os.makedirs(dpath, exist_ok=True); created.append(dname)
            else:
                existed.append(dname)
            # 00-报告打印
            dname = f"00-{cname}-{sname}-报告打印"
            dpath = os.path.join(root, dname)
            if not os.path.exists(dpath):
                os.makedirs(dpath, exist_ok=True); created.append(dname)
            else:
                existed.append(dname)
            # 02-保密承诺书
            nda_name = f"02-{cname}-{sname}-保密承诺书.docx"
            nda_path = os.path.join(root, nda_name)
            if os.path.exists(nda_path):
                existed.append(nda_name)
            else:
                try:
                    from utils.config import Config
                    tpl = os.path.join(Config.get_data_dir(), "templates", "02-保密承诺书模板.docx")
                    if os.path.exists(tpl):
                        import shutil, docx
                        shutil.copy2(tpl, nda_path)
                        doc = docx.Document(nda_path)
                        company = self.project.company_name or ""
                        create_date = self.project.created_at[:10] if self.project.created_at else ""
                        # 替换公司名: 逐run保留格式
                        for p in doc.paragraphs:
                            for run in p.runs:
                                if "XX公司" in run.text or "xx公司" in run.text:
                                    run.text = run.text.replace("XX公司", company).replace("xx公司", company)
                                    break
                        # 清除 split 的 "XX"+"公司" run 对
                        for p in doc.paragraphs:
                            for j in range(len(p.runs) - 1):
                                if p.runs[j].text.strip() == "XX" and p.runs[j+1].text.strip() == "公司":
                                    p.runs[j].text = company; p.runs[j+1].text = ""
                        # 替换日期: 保留每个run格式, 仅替换"XX"
                        if create_date:
                            parts = create_date.split("-")
                            if len(parts) == 3:
                                y, m, d = parts[0], f"{int(parts[1]):02d}", f"{int(parts[2]):02d}"
                                for table in doc.tables:
                                    for row in table.rows:
                                        for cell in row.cells:
                                            for p in cell.paragraphs:
                                                rs = p.runs
                                                if len(rs) >= 6 and all(rs[k].text.strip() == v for k, v in [(0,"XX"),(2,"XX"),(4,"XX")]):
                                                    rs[0].text = y; rs[2].text = m; rs[4].text = d
                                                    break
                        doc.save(nda_path)
                        created.append(nda_name)
                except Exception:
                    pass
            # 弹窗报告
            lines = []
            if created:
                lines.append("--- 已创建 ---")
                lines.extend(f"  + {x}" for x in created)
            if existed:
                lines.append("--- 已存在 ---")
                lines.extend(f"  = {x}" for x in existed)
            messagebox.showinfo("初始化完成", "\n".join(lines) if lines else "无需初始化")
        except Exception as e:
            messagebox.showerror("错误", f"初始化失败: {e}")

    def _on_zip_click(self):
        """处理"打包过程文档"按钮点击 - 将项目过程文件压缩为 ZIP

        打包策略：
          1. 查找项目文件夹
          2. 创建 ZIP 文件（命名格式：{公司}-{系统}-过程文档.zip）
          3. 按预定义的关键词列表筛选需要打包的文件：
             - 保密承诺书、测评调研表、测评授权书、风险告知书
             - 项目计划书、测评方案、首次会议记录、测评现场记录表
             - 问题汇总、漏洞扫描报告、项目文档移交清单、末次会议记录
             - 服务情况评价表、报备表
          4. 特殊处理"渗透测试报告"目录：
             - 如果目录非空：递归打包所有文件（保持目录结构）
             - 如果目录为空：添加空目录条目
          5. 成功打包后弹窗提示；无可打包文件时删除空 ZIP

        排除项：
          - 测评报告-终稿（不入过程文档包）
          - 报告打印相关文件
          - 其他归档文件
        """
        import os, zipfile  # 文件系统和 ZIP 压缩模块
        from tkinter import messagebox  # 消息弹窗

        try:
            root = self._find_project_folder()  # 查找项目文件夹路径
            if not root or not os.path.isdir(root):  # 文件夹不存在
                messagebox.showinfo("提示", "未找到项目文件夹")
                return

            # 构建 ZIP 文件名
            cname = self.project.company_name or "未命名"  # 公司名（取不到用"未命名"）
            sname = self.project.system_name or ""  # 系统名
            zip_name = f"{cname}-{sname}-过程文档.zip"  # ZIP 文件名格式
            zip_path = os.path.join(root, zip_name)  # ZIP 文件完整路径

            # ---- 需要打包的文件关键词列表 ----
            # 文件名中包含这些关键词之一的文件将被包含在 ZIP 中
            pack_keywords = [  # 仅打包 #3-#7 和 #9-#15
                "测评调研表", "测评授权书", "风险告知书",
                "项目计划书", "测评方案",
                "首次会议记录", "测评现场记录表",
                "问题汇总", "漏洞扫描报告",
                "项目文档移交清单", "末次会议记录",
            ]

            count = 0  # 已打包文件/目录计数
            with zipfile.ZipFile(zip_path, "w", zipfile.ZIP_DEFLATED) as zf:  # 创建 ZIP（DEFLATED 压缩）
                # --- 第一步：打包匹配关键词的单文件 ---
                for fname in os.listdir(root):  # 遍历根目录
                    fpath = os.path.join(root, fname)
                    if not os.path.isfile(fpath) or fname == zip_name:  # 跳过目录和正在创建的 ZIP
                        continue
                    name_no_ext = os.path.splitext(fname)[0]  # 取文件名（不含扩展名）
                    for kw in pack_keywords:  # 遍历关键词
                        if kw in name_no_ext:  # 文件名匹配关键词
                            zf.write(fpath, fname)  # 写入 ZIP（保持原文件名）
                            count += 1  # 计数 +1
                            break  # 匹配到一个关键词即处理下一个文件

                # --- 第二步：打包渗透测试报告目录（含所有内容） ---
                for dname in os.listdir(root):
                    dpath = os.path.join(root, dname)
                    if os.path.isdir(dpath) and "渗透测试报告" in dname:  # 是目标目录
                        has_files = False  # 目录是否有文件（非空标志）
                        for dirpath, _, filenames in os.walk(dpath):  # 递归遍历目录
                            for fn in filenames:  # 遍历每个文件
                                fp = os.path.join(dirpath, fn)  # 文件完整路径
                                arcname = os.path.relpath(fp, root).replace("\\", "/")  # ZIP 内路径（统一斜杠）
                                zf.write(fp, arcname)  # 写入 ZIP，保持目录结构
                                count += 1
                                has_files = True  # 标记为非空
                        # 如果目录为空，添加空目录条目（保留目录结构）
                        if not has_files:
                            info = zipfile.ZipInfo(dname + "/")  # 创建目录条目（末尾 / 标记为目录）
                            zf.writestr(info, "")  # 写入空内容
                            count += 1

            # ---- 结果处理 ----
            if count > 0:  # 至少打包了一个文件
                messagebox.showinfo("打包完成",
                    f"已打包 {count} 个文件\n{zip_name}")  # 显示打包结果
            else:  # 没有匹配的文件
                os.remove(zip_path)  # 删除空 ZIP 文件
                messagebox.showinfo("提示", "未找到可打包的过程文件")

        except Exception as e:  # 捕获所有异常
            messagebox.showerror("错误", f"打包失败: {e}")  # 弹窗显示错误

    # ==================================================================================
    # _on_report_print_click - 报告打印功能
    # ==================================================================================

    def _on_report_print_click(self):
        """处理"报告打印"按钮点击 - 生成测评报告打印信息并复制相关文件

        完整流程：
          1. 查找项目文件夹（如果找不到则提示并退出）
          2. 弹出 _show_report_dialog 编辑确认框，预填项目数据
          3. 用户确认 14 个字段后，查找/创建报告打印子目录
          4. 调用 _create_report_xlsx_data 生成测评报告打印信息.xlsx
          5. 复制相关文件到报告打印目录：
             - "测评授权书"相关文件
             - "风险告知书"相关文件
             - "测评报告-终稿.pdf"（如果有）
             - 过程文档 ZIP（如果有）
          6. 弹窗显示生成结果

        报告打印目录命名：00-{公司}-{系统}-报告打印
        XLSX 文件命名：00-{公司}-{系统}-测评报告打印信息.xlsx
        """
        import os, shutil  # 文件系统操作和高级文件复制（保留元数据）
        from tkinter import messagebox  # 消息弹窗
        from datetime import date  # 当前日期

        try:
            proot = self._find_project_folder()  # 查找项目文件夹根目录
            if not proot or not os.path.isdir(proot):  # 根目录不存在
                messagebox.showinfo("提示", "未找到项目文件夹")
                return

            # ---- 步骤 1: 弹出报告打印编辑确认框 ----
            data = _show_report_dialog(
                self,  # 父窗口
                cname=self.project.company_name or "",  # 预填公司名称
                sname=self.project.system_name or "",  # 预填系统名称
                location=(self.project.location or "").split("-")[0]  # 预填所属地（取省/市前缀）
                    if self.project.location else "",  # 无属地则空
                deadline=self.project.deadline or date.today().strftime("%Y-%m-%d"),  # 预填日期
            )
            if not data:  # 用户点击了取消
                return

            # 提取用户确认后的关键字段
            cname = data["cname"]  # 客户公司全称
            sname = data["sname"]  # 系统名称
            prefix = f"{cname}-{sname}"  # 文件名前缀

            # ---- 步骤 2: 查找或创建报告打印目录 ----
            report_dir = None  # 报告打印目录路径
            for dname in os.listdir(proot):  # 在项目根目录中查找
                if "报告打印" in dname and os.path.isdir(os.path.join(proot, dname)):
                    report_dir = os.path.join(proot, dname)  # 找到已存在的报告打印目录
                    break
            if not report_dir:  # 不存在则创建
                report_dir = os.path.join(proot, f"00-{prefix}-报告打印")  # 按标准格式命名
                os.makedirs(report_dir, exist_ok=True)  # 递归创建

            # ---- 步骤 3: 创建测评报告打印信息 XLSX ----
            xlsx_name = f"00-{prefix}-测评报告打印信息.xlsx"  # XLSX 文件名
            xlsx_path = os.path.join(report_dir, xlsx_name)  # XLSX 完整路径
            self._create_report_xlsx_data(xlsx_path, data, report_dir, proot)  # 生成 XLSX

            # ---- 步骤 4: 复制相关文件到报告打印目录 ----
            copied = 0  # 已复制文件计数

            # 复制"测评授权书"和"风险告知书"相关文件
            copy_keywords = ["测评授权书", "风险告知书"]
            for fname in os.listdir(proot):  # 遍历项目根目录
                fpath = os.path.join(proot, fname)
                if not os.path.isfile(fpath):  # 跳过目录
                    continue
                for kw in copy_keywords:  # 遍历复制关键词
                    if kw in fname:  # 文件名匹配关键词
                        shutil.copy2(fpath, os.path.join(report_dir, fname))  # 复制（保留元数据）
                        copied += 1  # 计数
                        break  # 匹配一个即处理下一个文件

                # 复制测评报告终稿 PDF
                if "测评报告-终稿" in fname and fname.lower().endswith(".pdf"):  # 必须是 PDF
                    shutil.copy2(fpath, os.path.join(report_dir, fname))  # 复制到打印目录
                    copied += 1

            # 复制过程文档 ZIP（如果已生成）
            zip_name = f"{cname}-{sname}-过程文档.zip"  # ZIP 文件名
            zip_src = os.path.join(proot, zip_name)  # ZIP 源路径
            if os.path.exists(zip_src):  # ZIP 文件存在
                shutil.copy2(zip_src, os.path.join(report_dir, zip_name))  # 复制
                copied += 1

            # ---- 步骤 5: 显示结果 ----
            messagebox.showinfo("报告打印完成",
                f"已生成 {xlsx_name}\n已复制 {copied} 个文件到报告打印目录")

        except Exception as e:  # 捕获所有异常
            messagebox.showerror("错误", f"报告打印失败: {e}")  # 显示错误

    # ==================================================================================
    # _create_report_xlsx_data / _create_report_xlsx - XLSX 生成
    # ==================================================================================

    def _create_report_xlsx_data(self, path, data, report_dir, root):
        """根据编辑框提交的数据创建测评报告打印信息 XLSX 文件

        这是两步 XLSX 生成流程的入口：
          1. 先调用 _create_report_xlsx 创建基础 XLSX（含表头、固定数据行）
          2. 再用编辑框的额外数据覆盖 B3、D3、J3~U3 等单元格

        Args:
            path: XLSX 文件保存路径
            data: 用户确认的 14 个字段字典（来自 _show_report_dialog 返回值）
            report_dir: 报告打印目录路径（用于附件超链接）
            root: 项目根目录路径（用于附件超链接查找）
        """
        # 第一步：创建基础 XLSX（表头 + 固定列数据 + 附件超链接）
        self._create_report_xlsx(path, data["cname"], data["sname"], report_dir, root)

        # 第二步：用编辑框数据二次写入（覆盖用户编辑的字段）
        import openpyxl  # Excel 文件操作库
        wb = openpyxl.load_workbook(path)  # 加载刚创建的基础 XLSX
        ws = wb.active  # 获取活动工作表
        from openpyxl.styles import Alignment, Border, Side  # Excel 样式

        # 定义细线边框（统一所有数据单元格的边框样式）
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),  # 左右细线
            top=Side(style='thin'), bottom=Side(style='thin'))  # 上下细线

        # 编辑框额外数据与单元格的映射关系
        extra = {
            'B3': data.get("contract", ""),  # 合同编号或项目名称 -> B 列第 3 行
            'D3': data.get("crm", "是"),  # 是否录入 CRM -> D 列第 3 行
            'J3': data.get("author", ""),  # 编制人 -> J 列第 3 行
            'K3': data.get("reviewer", ""),  # 审核人 -> K 列第 3 行
            'L3': data.get("pentester", ""),  # 渗透人员 -> L 列第 3 行
            'M3': data.get("conclusion", ""),  # 测评结论 -> M 列第 3 行
            'N3': data.get("seal", ""),  # 盖章 -> N 列第 3 行
            'S3': data.get("print_req", ""),  # 打印要求 -> S 列第 3 行
            'T3': data.get("leader", ""),  # 项目组长 -> T 列第 3 行
            'U3': data.get("actual_author", ""),  # 实际编制人 -> U 列第 3 行
        }
        for ref, val in extra.items():  # 遍历映射写入数据
            cell = ws[ref]  # 获取单元格
            cell.value = val  # 写入值
            cell.border = thin_border  # 设置边框
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # 设置对齐

        wb.save(path)  # 保存 XLSX 文件

    def _create_report_xlsx(self, path, cname, sname, report_dir, root):
        """创建测评报告打印信息 XLSX 的基础结构

        生成一个符合测评报告打印规范的 Excel 文件，包含：

        表结构：
          - 行 1 (G1)：提示文字 "(需要签入报告，请确认)"（红色粗体）
          - 行 2：表头行（21 列，蓝底白字粗体）
          - 行 3：数据行（项目实际数据）

        表头列（A~U 共 21 列）：
          A: 序号 | B: 合同编号 | C: 客户公司 | D: 是否录入CRM | E: 所属地
          F: 系统名称 | G: 编制日期 | H: 审核日期 | I: 批准日期
          J: 编制人 | K: 审核人 | L: 渗透人员
          M: 测评结论数量 | N: 盖章
          O: 基本情况表附件 | P: 授权书附件 | Q: 报告附件 | R: 归档附件
          S: 打印要求 | T: 项目组长 | U: 实际编制人

        附件超链接（O~R 列）：
          O: 附件"基本情况表" -> 超链接到对应文件
          P: 附件"测评授权书" + "风险告知书" -> 超链接到对应文件
          Q: 附件"测评报告-终稿" -> 超链接到对应 PDF
          R: 附件"过程文档.zip" -> 超链接到 ZIP 文件

        Args:
            path: XLSX 文件保存路径
            cname: 客户公司全称
            sname: 系统名称
            report_dir: 报告打印目录（优先搜索附件）
            root: 项目根目录（兜底搜索附件）
        """
        import os as _os  # 操作系统模块（使用别名避免与外部 os 冲突）
        import openpyxl  # Excel 文件操作库
        from openpyxl.styles import Font, Alignment, Border, Side, PatternFill  # Excel 样式组件
        from datetime import date  # 当前日期

        # ---- 创建 Workbook 和默认工作表 ----
        wb = openpyxl.Workbook()  # 创建新的 Excel 工作簿
        ws = wb.active  # 获取默认的活动工作表
        ws.title = "Sheet1"  # 设置工作表名称为 "Sheet1"

        # ---- 设置列宽（A~U 共 21 列） ----
        widths = {'A': 6, 'B': 30, 'C': 22, 'D': 10, 'E': 10, 'F': 18,
                  'G': 12, 'H': 12, 'I': 12, 'J': 12, 'K': 10, 'L': 10,
                  'M': 16, 'N': 10, 'O': 16, 'P': 16, 'Q': 14, 'R': 14,
                  'S': 16, 'T': 10, 'U': 14}
        for col, w in widths.items():  # 遍历列宽设置
            ws.column_dimensions[col].width = w  # 设置每列宽度

        # ---- 提示行 (第 1 行) ----
        ws['G1'] = '（需要签入报告，请确认）'  # 红色提醒文字
        ws['G1'].font = Font(color='FF0000', bold=True)  # 红色粗体

        # ---- 表头 (第 2 行) ----
        headers = ['序号', '合同编号或项目名称', '客户公司全称', '是否录入CRM', '所属地',
                   '系统名称', '编制日期', '审核日期', '批准日期',
                   '编制人（与联盟系统对应）', '审核人', '渗透人员',
                   '测评结论及重大风险隐患数量', '盖章',
                   '等级测评项目基本情况表附件', '授权书及风险告知书附件',
                   '测评报告附件', '测评归档附件', '打印要求',
                   '项目组长联系人', '实际报告编制人']  # 21 个表头

        # 表头样式定义
        header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')  # 浅蓝填充
        thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),  # 左右细线
            top=Side(style='thin'), bottom=Side(style='thin'))  # 上下细线

        for i, h in enumerate(headers, 1):  # 遍历表头，i 从 1（Excel 列号从 1 开始）
            cell = ws.cell(row=2, column=i, value=h)  # 在第 2 行写入表头
            cell.font = Font(bold=True, size=10)  # 粗体 10 号
            cell.fill = header_fill  # 浅蓝背景填充
            cell.border = thin_border  # 细线边框
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # 居中自动换行
        ws.row_dimensions[2].height = 35  # 表头行高 35px

        # ---- 数据行 (第 3 行，紧接表头无空行) ----
        date_val = self.project.deadline or date.today().strftime('%Y-%m-%d')  # 日期取截止日或当天
        data = [
            (1, 'A'),  # 序号：固定为 1
            (f'{cname}网络安全等级保护测评服务项目', 'B'),  # 合同编号
            (cname, 'C'),  # 客户公司全称
            ('是', 'D'),  # 是否录入 CRM（默认"是"）
            ((self.project.location or '').split('-')[0] if self.project.location else '', 'E'),  # 所属地
            (sname, 'F'),  # 系统名称
            (date_val, 'G'), (date_val, 'H'), (date_val, 'I'),  # 编制/审核/批准日期（统一）
            ('双击打开附件', 'O'), ('双击打开附件', 'P'),  # 附件提示
            ('双击打开附件', 'Q'), ('双击打开附件', 'R'),  # 附件提示
        ]
        for val, col in data:  # 遍历写入每个数据项
            cell = ws[f'{col}3']  # 定位到第 3 行对应列
            cell.value = val  # 设置值
            cell.border = thin_border  # 设置边框
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)  # 居中

        # 为所有 A~U 列第 3 行设置完整边框（包括编辑框数据将覆盖的单元格）
        for col_idx in range(1, 22):  # Excel 列 1~21（A~U）
            cell = ws.cell(row=3, column=col_idx)  # 获取第 3 行该列的单元格
            cell.border = thin_border  # 设置边框
            if not cell.value:  # 空单元格也设置对齐
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        ws.row_dimensions[3].height = 25  # 数据行高 25px

        # ---- 合并单元格 G1-U1（提示行合并为一整行） ----
        ws.merge_cells('G1:U1')  # 将 G1 到 U1 合并为一个单元格

        # ---- 附件超链接设置 ----
        # O~R 列：分别对应基本情况表、授权书+风险告知书、测评报告、过程文档
        link_map = {
            'O': ['基本情况表'],  # O 列 -> 查找包含"基本情况表"的文件
            'P': ['测评授权书', '风险告知书'],  # P 列 -> 查找授权书或风险告知书
            'Q': ['测评报告-终稿'],  # Q 列 -> 查找终稿 PDF
            'R': ['过程文档.zip'],  # R 列 -> 查找过程文档 ZIP
        }

        for col, keywords in link_map.items():  # 遍历每个附件列
            found = False  # 是否已找到并设置超链接
            for kw in keywords:  # 遍历该列的关键词
                for d in [report_dir, root]:  # 先查报告打印目录，再查项目根目录
                    if not _os.path.isdir(d):  # 目录不存在
                        continue
                    for fname in _os.listdir(d):  # 遍历目录中的所有文件
                        if kw in fname:  # 文件名匹配关键词
                            fpath = _os.path.join(d, fname)  # 拼接完整路径
                            cell = ws[f'{col}3']  # 获取对应列的单元格
                            cell.value = fname  # 设置单元格显示为文件名
                            cell.hyperlink = fpath  # 设置为文件超链接
                            cell.font = Font(color='0563C1', underline='single', size=10)  # 蓝色下划线样式
                            found = True  # 标记已找到
                            break  # 跳出文件遍历
                    if found:  # 已找到
                        break  # 跳出目录遍历
                if found:  # 已找到
                    break  # 跳出关键词遍历

        # ---- 保存文件 ----
        wb.save(path)  # 保存 XLSX 文件到指定路径

    # ==================================================================================
    # 事件绑定 - 鼠标交互事件的递归绑定
    # ==================================================================================

    def _bind_events(self):
        """为卡片内所有子组件递归绑定鼠标事件

        设计原则：
          - 直接子组件不包括按钮（按钮通过 command 独立处理点击）
          - 递归绑定到每个子组件及其后代，实现整张卡片的统一交互体验
          - 按钮组件被排除，避免干扰其自身的 command 事件

        绑定的事件：
          - <Button-1>：单击 -> 选中/取消选中卡片
          - <Double-Button-1>：双击 -> 打开编辑对话框
          - <Enter>：鼠标进入 -> 切换悬停背景色
          - <Leave>：鼠标离开 -> 恢复默认背景色

        排除的按钮集合（btn_widgets）包含：
          左右箭头按钮、详情/编辑/复制按钮、文件夹/重命名/打包/报告按钮
        """
        btn_widgets = {self._prev_btn, self._next_btn,  # 左右箭头按钮
                       self._detail_btn, self._edit_btn, self._copy_btn,  # 操作按钮
                       self._folder_btn, self._init_btn,  # 文件夹+初始化
                       self._rename_btn, self._zip_btn,  # 重命名+打包
                       self._report_btn}  # 报告打印按钮

        # 先给 Frame 自身绑定事件（捕获卡片边缘区域的鼠标事件）
        self.bind("<Button-1>", self._on_click)  # 鼠标左键单击
        self.bind("<Double-Button-1>", self._on_double_click)  # 鼠标左键双击
        self.bind("<Enter>", self._on_enter)  # 鼠标进入组件区域
        self.bind("<Leave>", self._on_leave)  # 鼠标离开组件区域

        def _bind_recursive(widget):
            """递归函数：遍历组件树，为除按钮外的所有组件绑定鼠标事件

            Args:
                widget: 当前要处理的 Tkinter 组件
            """
            if widget not in btn_widgets:  # 跳过按钮组件（它们有自己的 command 绑定）
                widget.bind("<Button-1>", self._on_click)  # 单击
                widget.bind("<Double-Button-1>", self._on_double_click)  # 双击
                widget.bind("<Enter>", self._on_enter)  # 进入
                widget.bind("<Leave>", self._on_leave)  # 离开
                for child in widget.winfo_children():  # 递归遍历该组件的所有子组件
                    _bind_recursive(child)  # 继续递归

        # 从 Frame 的第一层子组件开始递归绑定（不包括 Frame 自身）
        for child in self.winfo_children():  # 遍历 Frame 的直接子组件
            _bind_recursive(child)  # 进入递归绑定

    # ==================================================================================
    # 鼠标事件处理方法
    # ==================================================================================

    def _on_click(self, event):
        """处理鼠标单击事件 - 转发给外部回调（选中/取消选中）"""
        if self.on_click:
            self.on_click(self)

    def _on_double_click(self, event):
        """处理鼠标双击事件 - 转发给外部回调（打开编辑对话框）"""
        if self.on_double_click:
            self.on_double_click(self)

    def _on_enter(self, event):
        """处理鼠标进入卡片区域 - 切换为悬停背景色

        仅在卡片未选中状态下生效，选中状态不受悬浮影响。
        """
        if not self.is_selected:  # 仅在未选中时生效
            self.configure(bg=Config.CARD_HOVER_BG)  # 设置 Frame 自身背景
            self._set_bg_recursive(self, Config.CARD_HOVER_BG)  # 递归设置所有子组件背景

    def _on_leave(self, event):
        """处理鼠标离开卡片区域 - 恢复默认背景色

        仅在卡片未选中状态下生效，选中状态保持蓝色边框。
        """
        if not self.is_selected:  # 仅在未选中时生效
            self.configure(bg=Config.CARD_BG)  # 恢复 Frame 默认白色背景
            self._set_bg_recursive(self, Config.CARD_BG)  # 递归恢复所有子组件背景

    def _set_bg_recursive(self, widget, color):
        """递归设置组件树的背景色

        遍历整个组件树，将符合条件的组件的背景色统一设置为指定颜色。
        排除以下固定颜色组件（避免覆盖功能按钮的设计颜色）：
          - 主题色按钮：蓝色(#3498db, #2980b9)、绿色(#27ae60, #1abc9c)
          - 状态颜色：橙色(#e67e22, #f39c12)、紫色(#8e44ad, #9b59b6)
          - 红色(#e74c3c)、浅灰(#ecf0f1, #d5dbdb)、灰色(#b0b8c1, #95a5a6)
          - 白色(white)：输入框等

        Args:
            widget: 要处理的根组件
            color: 目标背景色
        """
        try:
            bg = widget.cget("bg")  # 获取组件当前背景色（可能抛出 TclError）
            # 检查当前背景色是否在固定颜色排除列表中
            if bg not in ("#3498db", "#2ecc71", "#e67e22",
                          "#9b59b6", "#e74c3c", "#1abc9c",
                          "#f39c12", "#95a5a6",
                          "#ecf0f1", "#d5dbdb", "#2980b9",
                          "#b0b8c1", "white"):
                widget.configure(bg=color)  # 设置新背景色
        except tk.TclError:  # 某些组件可能不支持 bg 属性（如 Canvas 等的特殊子项）
            pass  # 忽略错误，继续处理其他组件

    # ==================================================================================
    # 选中状态管理
    # ==================================================================================

    def set_selected(self, selected: bool):
        """设置卡片的选中状态

        选中状态效果：
          - 选中 (True)：蓝色 (#2196F3) 2px 加粗边框
          - 取消 (False)：恢复默认浅灰 (#d0d5dd) 1px 边框和白色背景

        Args:
            selected: True = 选中，False = 取消选中
        """
        self.is_selected = selected  # 更新内部选中状态标志
        if selected:  # 设置为选中
            # 蓝色材质风格边框，2px 加粗
            self.configure(highlightbackground="#2196F3",
                           highlightthickness=2)
        else:  # 取消选中
            # 恢复默认边框颜色和宽度
            self.configure(highlightbackground=Config.CARD_BORDER,
                           highlightthickness=1)
            self.configure(bg=Config.CARD_BG)  # 恢复默认白色背景

    def refresh(self):
        """刷新卡片显示（销毁所有子组件并重建 UI）

        当项目数据发生修改后调用，通过销毁并重建的方式更新卡片显示内容。
        重建后自动恢复所有回调函数引用，保证交互功能不丢失。

        使用场景：
          - 项目数据被外部修改（编辑对话框、详情窗口）
          - 需要刷新卡片显示以反映最新数据
        """
        # 保存当前所有回调函数引用（重建后需要恢复）
        saved = {
            "on_click": self.on_click,  # 单击回调
            "on_double_click": self.on_double_click,  # 双击回调
            "on_detail": self.on_detail,  # 详情按钮回调
            "on_edit": self.on_edit,  # 编辑按钮回调
            "on_copy": self.on_copy,  # 复制按钮回调
            "on_move_prev": self.on_move_prev,  # 左箭头回调
            "on_move_next": self.on_move_next,  # 右箭头回调
        }
        for widget in self.winfo_children():  # 销毁所有子组件（释放 Tkinter 资源）
            widget.destroy()
        self._build_ui()  # 重新构建 UI（根据最新的 project 数据）
        self._bind_events()  # 重新绑定鼠标事件
        for k, v in saved.items():  # 恢复保存的回调函数
            setattr(self, k, v)  # 使用 setattr 动态设置属性
